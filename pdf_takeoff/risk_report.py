"""Relatório de risco/confiança do quantitativo, em português e inglês.

A ideia: em vez de o usuário reconferir a planta inteira, o programa aponta
exatamente onde a chance de erro é maior — e ordena pelo impacto em $, não
pela quantidade de itens — para a revisão manual focar no que mais pesa no
orçamento. Isso é tudo determinístico (geometria + regras), não depende de
nenhuma chamada de IA.

Duas fontes de risco cobertas aqui:
1. Segmentos de parede que não puderam ser confirmados por um keynote de
   texto perto o suficiente (`MatchedSegment.source != "keynote"`), ou cujo
   keynote mais próximo diverge do que a legenda de cor sugere.
2. Blocos de quantidade "suspeitosamente idênticos" entre partes do projeto
   que deveriam ser diferentes (ex.: teto do térreo do prédio 2 igual ao
   térreo do prédio 1, apesar de áreas de piso bem diferentes) — o tipo de
   erro de copiar/colar que apareceu na conferência do projeto Clevedon.
"""

from __future__ import annotations

from dataclasses import dataclass, field

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .gib_spec_extract import BoardSpec
from .wall_linings_plan import MatchedSegment

_SEVERITY_ORDER = {"high": 0, "medium": 1, "low": 2}
_SEVERITY_LABEL = {
    "high": ("ALTO", "HIGH"),
    "medium": ("MÉDIO", "MEDIUM"),
    "low": ("BAIXO", "LOW"),
}
_SEVERITY_FILL = {
    "high": PatternFill("solid", fgColor="FADBD8"),
    "medium": PatternFill("solid", fgColor="FCF3CF"),
    "low": PatternFill("solid", fgColor="EAECEE"),
}


@dataclass
class Finding:
    severity: str  # "high" | "medium" | "low"
    title_pt: str
    title_en: str
    detail_pt: str
    detail_en: str
    dollar_impact: float | None = None
    location: str = ""


def _rate_for_spec(spec: BoardSpec, rate_lookup: dict[str, float]) -> float | None:
    rate = rate_lookup.get(spec.board_type_label)
    if rate is None:
        return None
    return rate * spec.layers


def find_wall_segment_risks(
    matched: list[MatchedSegment],
    rate_lookup: dict[str, float] | None = None,
    wall_height_m: float = 3.0,
    sheet_label: str = "",
) -> list[Finding]:
    """Agrupa os segmentos de parede não confirmados por keynote (cor só, ou
    cor não mapeada) e os que tiveram divergência cor x keynote, um Finding
    por grupo — com o $ em risco estimado, se uma tabela de taxas for dada."""
    rate_lookup = rate_lookup or {}
    findings: list[Finding] = []

    by_color_only: dict[tuple, float] = {}
    by_unmapped: dict[tuple, float] = {}
    # agrupado por keynote_code (nao um Finding por segmento) -- numa planta
    # densa (varias salas pequenas perto uma da outra), o "vizinho mais
    # proximo" as vezes e' na verdade a tag de uma parede adjacente, nao a
    # dela mesma; isso tende a se repetir dezenas de vezes pro MESMO codigo,
    # entao agrupar por codigo e' o que da um sinal util em vez de ruido.
    mismatches_by_code: dict[str, list[MatchedSegment]] = {}

    for m in matched:
        if m.source == "color_only":
            by_color_only[m.color] = by_color_only.get(m.color, 0.0) + m.length_m
        elif m.source == "unmapped_color":
            by_unmapped[m.color] = by_unmapped.get(m.color, 0.0) + m.length_m
        if m.color_spec_mismatch and m.keynote_code:
            mismatches_by_code.setdefault(m.keynote_code, []).append(m)

    for color, length_m in by_color_only.items():
        spec = None
        for m in matched:
            if m.color == color and m.board_spec is not None:
                spec = m.board_spec
                break
        dollar = None
        if spec is not None:
            rate = _rate_for_spec(spec, rate_lookup)
            if rate is not None:
                dollar = length_m * wall_height_m * rate
        label = spec.board_type_label if spec else f"cor {color}"
        findings.append(
            Finding(
                severity="medium",
                title_pt=f"{length_m:.1f} m de '{label}' sem keynote próximo confirmando camadas/espessura",
                title_en=f"{length_m:.1f} m of '{label}' with no nearby keynote confirming layers/thickness",
                detail_pt=(
                    f"Esses {length_m:.1f} m de parede foram classificados só pela cor da linha "
                    f"(legenda 'WALL LININGS KEY'), sem um código de keynote perto o suficiente para "
                    f"confirmar quantas camadas ou a espessura exata. Foi assumido 1 camada. "
                    f"Se parte disso for na verdade de 2 camadas, o custo real de chapa é maior."
                ),
                detail_en=(
                    f"These {length_m:.1f} m of wall were classified by line colour only (the "
                    f"'WALL LININGS KEY' legend), with no keynote tag close enough to confirm layer "
                    f"count or exact thickness. Single layer was assumed. If part of this is actually "
                    f"double-layer, real board cost is higher."
                ),
                dollar_impact=dollar,
                location=sheet_label,
            )
        )

    for color, length_m in by_unmapped.items():
        findings.append(
            Finding(
                severity="high",
                title_pt=f"{length_m:.1f} m de parede com cor não reconhecida na legenda",
                title_en=f"{length_m:.1f} m of wall with a colour not found in the legend",
                detail_pt=(
                    f"Existe {length_m:.1f} m de linha de forro nesta prancha com uma cor que não "
                    f"bate com nenhuma entrada da 'WALL LININGS KEY'. Pode ser uma revisão de legenda "
                    f"não capturada, ou uma cor usada para outra coisa (dimensão, hachura). Ficou de "
                    f"fora do quantitativo — confira manualmente."
                ),
                detail_en=(
                    f"There is {length_m:.1f} m of lining line on this sheet with a colour that does "
                    f"not match any entry in the 'WALL LININGS KEY'. Could be a legend revision not "
                    f"captured, or a colour used for something else (dimension, hatch). Left out of "
                    f"the takeoff — please check manually."
                ),
                dollar_impact=None,
                location=sheet_label,
            )
        )

    for code, segs in mismatches_by_code.items():
        total_len = sum(s.length_m for s in segs)
        board_label = segs[0].board_spec.board_type_label if segs[0].board_spec else "?"
        avg_dist = sum(s.keynote_dist_m or 0 for s in segs) / len(segs)
        # muitos segmentos curtos batendo no mesmo keynote de longe = provavel
        # tag de parede vizinha, nao erro de desenho -> risco medio, nao alto.
        severity = "high" if len(segs) <= 3 or total_len >= 10 else "medium"
        findings.append(
            Finding(
                severity=severity,
                title_pt=f"Keynote {code} não bate com a cor em {len(segs)} trecho(s), {total_len:.1f} m no total",
                title_en=f"Keynote {code} disagrees with colour on {len(segs)} segment(s), {total_len:.1f} m total",
                detail_pt=(
                    f"Em {len(segs)} trecho(s) de parede (somando {total_len:.1f} m), o keynote mais "
                    f"próximo foi '{code}' ({board_label}), mas a cor da linha na planta sugere outro "
                    f"produto. Distância média até a tag: {avg_dist:.1f} m. Se forem muitos trechos "
                    f"curtos e espalhados, provavelmente a tag mais próxima é de uma parede vizinha "
                    f"(comum em salas pequenas e próximas) — não necessariamente um erro. Vale conferir "
                    f"visualmente na planta antes de confiar na especificação."
                ),
                detail_en=(
                    f"On {len(segs)} wall segment(s) (totalling {total_len:.1f} m), the nearest keynote "
                    f"was '{code}' ({board_label}), but the line colour on the plan suggests a different "
                    f"product. Average distance to the tag: {avg_dist:.1f} m. If these are many short, "
                    f"scattered segments, the nearest tag is probably for a neighbouring wall (common in "
                    f"small, closely-spaced rooms) — not necessarily a drawing error. Worth a visual check "
                    f"before trusting the spec."
                ),
                dollar_impact=None,
                location=sheet_label,
            )
        )

    return findings


def detect_duplicate_blocks(
    named_groups: dict[str, list[float]],
    tolerance: float = 1e-6,
) -> list[Finding]:
    """Compara os vetores de valores de cada grupo nomeado (ex.: quantidades
    de teto de cada piso) dois a dois; sinaliza quando dois grupos DIFERENTES
    têm valores idênticos (dentro da tolerância) em TODOS os itens — indício
    forte de copiar/colar sem atualizar, não coincidência."""
    findings: list[Finding] = []
    names = list(named_groups)
    for i in range(len(names)):
        for j in range(i + 1, len(names)):
            a, b = named_groups[names[i]], named_groups[names[j]]
            if not a or not b or len(a) != len(b):
                continue
            if all(abs(x - y) <= tolerance for x, y in zip(a, b)) and any(abs(x) > tolerance for x in a):
                findings.append(
                    Finding(
                        severity="high",
                        title_pt=f"'{names[i]}' e '{names[j]}' têm exatamente os mesmos valores",
                        title_en=f"'{names[i]}' and '{names[j]}' have exactly the same values",
                        detail_pt=(
                            f"Todos os {len(a)} valores de '{names[i]}' e '{names[j]}' são idênticos, "
                            f"casa decimal por casa decimal. Para duas partes do projeto que deveriam "
                            f"ter geometria diferente, isso é estatisticamente muito improvável — forte "
                            f"indício de que um bloco foi copiado e colado sem atualizar os números. "
                            f"Confirme com quem gerou os dados antes de orçar."
                        ),
                        detail_en=(
                            f"All {len(a)} values in '{names[i]}' and '{names[j]}' are identical down "
                            f"to the decimal. For two parts of the project that should have different "
                            f"geometry, this is statistically very unlikely — a strong sign that a block "
                            f"was copy-pasted without updating the numbers. Confirm with whoever produced "
                            f"the data before pricing."
                        ),
                        dollar_impact=None,
                        location=f"{names[i]} / {names[j]}",
                    )
                )
    return findings


# --------------------------- planilha bilíngue -------------------------------

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(color="FFFFFF", bold=True)


def _style_header(ws: Worksheet, row: int, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_bilingual_risk_report(
    findings: list[Finding],
    output_path: str,
    quantities: dict[str, dict[BoardSpec, float]] | None = None,
) -> None:
    """Gera um .xlsx com uma aba 'Findings / Achados' (PT e EN lado a lado,
    ordenado por severidade e depois por $ em risco) e, se `quantities` for
    passado ({rótulo_da_planta: {BoardSpec: metros}}), uma aba de
    quantidades também bilíngue, com a camada (1x/2x) já separada."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Findings - Achados"

    ws["A1"] = "RISK REPORT — REVIEW THESE FIRST  /  RELATÓRIO DE RISCO — REVISE ISTO PRIMEIRO"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = "Ordenado por severidade e depois por impacto em $ estimado, do maior para o menor."
    ws["A2"].font = Font(italic=True, size=9, color="666666")

    headers = [
        "Severity / Severidade",
        "Location / Local",
        "Title (EN)",
        "Título (PT)",
        "Detail (EN)",
        "Detalhe (PT)",
        "Est. $ at risk / $ em risco (est.)",
    ]
    r = 4
    for i, h in enumerate(headers, start=1):
        ws.cell(row=r, column=i, value=h)
    _style_header(ws, r, len(headers))
    r += 1

    ordered = sorted(
        findings,
        key=lambda f: (_SEVERITY_ORDER.get(f.severity, 9), -(f.dollar_impact or 0)),
    )
    for f in ordered:
        sev_pt, sev_en = _SEVERITY_LABEL.get(f.severity, (f.severity, f.severity))
        ws.cell(row=r, column=1, value=f"{sev_en} / {sev_pt}")
        ws.cell(row=r, column=2, value=f.location)
        ws.cell(row=r, column=3, value=f.title_en)
        ws.cell(row=r, column=4, value=f.title_pt)
        ws.cell(row=r, column=5, value=f.detail_en).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=r, column=6, value=f.detail_pt).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=r, column=7, value=round(f.dollar_impact, 2) if f.dollar_impact is not None else None)
        fill = _SEVERITY_FILL.get(f.severity)
        if fill:
            for c in range(1, 8):
                ws.cell(row=r, column=c).fill = fill
        ws.row_dimensions[r].height = 60
        r += 1

    for col, width in ((1, 16), (2, 16), (3, 34), (4, 34), (5, 46), (6, 46), (7, 16)):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A5"

    if quantities:
        ws2 = wb.create_sheet("Quantities - Quantidades")
        headers2 = [
            "Sheet / Prancha", "Board type / Tipo de chapa", "Layers / Camadas",
            "Linear metres / Metros lineares",
        ]
        r = 1
        for i, h in enumerate(headers2, start=1):
            ws2.cell(row=r, column=i, value=h)
        _style_header(ws2, r, len(headers2))
        r += 1
        for sheet_label, totals in quantities.items():
            for spec, length_m in sorted(totals.items(), key=lambda x: -x[1]):
                ws2.cell(row=r, column=1, value=sheet_label)
                ws2.cell(row=r, column=2, value=spec.board_type_label)
                ws2.cell(row=r, column=3, value=spec.layers)
                ws2.cell(row=r, column=4, value=round(length_m, 2))
                r += 1
        for col, width in ((1, 20), (2, 20), (3, 12), (4, 20)):
            ws2.column_dimensions[get_column_letter(col)].width = width

    wb.save(output_path)
