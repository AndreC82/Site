"""Gera a planilha de quantidades/orçamento no formato usado pelo estimador:
agrupada por tipo de forro (Fireline, Aqualine, Standard...) e altura de parede,
com 3 taxas por linha (GIB, Stopping, Pintura) em $/m² puxadas de uma aba
"Taxas" única. O Gib é precificado por tipo/espessura (instalação + custo da
chapa dividido pela área da chapa), Pintura e Stopping são valores únicos.
Todos os valores de custo levam uma margem (%) até virarem preço de venda.
Uma aba Resumo soma GIB/Plaster/Painting/Total, Margem e Contingência.
"""

from __future__ import annotations

from dataclasses import dataclass, field

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
_INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_BOLD = Font(bold=True)

# Colunas fixas da aba "Quantities".
COL_DESC = 1  # A
COL_QTY = 2  # B
COL_QTY_UNIT = 3  # C
COL_HEIGHT = 4  # D
COL_TOTAL = 5  # E
COL_TOTAL_UNIT = 6  # F
COL_GIB_RATE = 8  # H
COL_GIB_TOTAL = 9  # I
COL_STOP_RATE = 11  # K
COL_STOP_TOTAL = 12  # L
COL_PAINT_RATE = 14  # N
COL_PAINT_TOTAL = 15  # O
COL_ROW_TOTAL = 17  # Q

# Colunas fixas da aba "Taxas".
TAXAS_COL_LABEL = 2  # B
TAXAS_COL_INSTALL = 3  # C (instalação/mão de obra, custo $/m² - só linhas de Gib)
TAXAS_COL_BOARD_COST = 4  # D (custo da chapa $ - só linhas de Gib)
TAXAS_COL_BOARD_COST_M2 = 5  # E (custo da chapa por m², fórmula - só linhas de Gib)
TAXAS_COL_COST_TOTAL = 6  # F (custo total $/m² ou $/unidade)
TAXAS_COL_SALE = 7  # G (venda = custo x (1+margem) - é isso que a Quantities referencia)

TAXAS_MARGIN_CELL = "'Taxas'!$C$4"
TAXAS_BOARD_WIDTH_CELL = "'Taxas'!$C$6"
TAXAS_BOARD_HEIGHT_CELL = "'Taxas'!$C$7"
TAXAS_BOARD_AREA_CELL = "'Taxas'!$C$8"

# Tipos de chapa de Gib padrão (linhas 11-18 da tabela de taxas).
GIB_BOARD_TYPES = [
    "10mm Standard",
    "10mm Aqualine",
    "13mm Standard",
    "13mm Aqualine",
    "13mm Fireline",
    "13mm Noiseline",
    "13mm Braceline",
    "16mm Fireline",
    "19mm Fireline",
]
_GIB_TABLE_FIRST_ROW = 11

# Itens de taxa única (linhas 20-27).
_PAINTING_ROW = 20
_STOPPING_WALL_ROW = 21
_STOPPING_CEILING_ROW = 22
_CORNER_ROW = 23
_SEALANT_ROW = 24
_SKIRTING_ROW = 25
_SINGLE_DOOR_ROW = 26
_DOUBLE_DOOR_ROW = 27


def _sale_cell(row: int) -> str:
    return f"'Taxas'!${get_column_letter(TAXAS_COL_SALE)}${row}"


TAXAS_PAINTING = _sale_cell(_PAINTING_ROW)
TAXAS_STOPPING_WALL = _sale_cell(_STOPPING_WALL_ROW)
TAXAS_STOPPING_CEILING = _sale_cell(_STOPPING_CEILING_ROW)
TAXAS_CORNER = _sale_cell(_CORNER_ROW)
TAXAS_SEALANT = _sale_cell(_SEALANT_ROW)
TAXAS_SKIRTING_PAINT = _sale_cell(_SKIRTING_ROW)
TAXAS_SINGLE_DOOR = _sale_cell(_SINGLE_DOOR_ROW)
TAXAS_DOUBLE_DOOR = _sale_cell(_DOUBLE_DOOR_ROW)


@dataclass
class GibBoardRate:
    board_type: str  # deve bater com um dos GIB_BOARD_TYPES
    install_cost_m2: float  # mão de obra, $/m² de custo
    board_cost: float  # custo da chapa inteira, $


@dataclass
class LiningItem:
    description: str  # ex.: "1x 10mm Standard", "2x 19mm Fireline"
    qty: float  # metros lineares de parede com esse tipo de forro
    board_type: str  # referencia um item de GIB_BOARD_TYPES
    qty_unit: str = "m"
    layers: int = 1  # 1 = single layer, 2 = double layer (dobra o custo de Gib)
    include_stopping: bool = True
    include_painting: bool = True
    total_unit: str = "m2"


@dataclass
class HeightGroup:
    name: str  # ex.: "LININGS - 2.70m"
    height_m: float
    items: list[LiningItem] = field(default_factory=list)
    corner_trims_qty: float = 0.0  # contagem de cantoneiras (ea)
    sealant_qty: float = 0.0  # metros lineares de selante de rodapé


@dataclass
class CeilingItem:
    description: str  # ex.: "13mm Standard Gib board (C1)"
    area_m2: float
    board_type: str
    include_stopping: bool = True
    include_painting: bool = True


@dataclass
class DoorPaintItem:
    description: str  # ex.: "D1 - Single Swing"
    count: int
    door_type: str = "single"  # "single" ou "double"


def _style_header(ws: Worksheet, row: int, cols: list[int]) -> None:
    for col in cols:
        cell = ws.cell(row=row, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _section_row(ws: Worksheet, row: int, text: str) -> None:
    cell = ws.cell(row=row, column=COL_DESC, value=text)
    cell.font = _BOLD
    for col in range(1, 18):
        ws.cell(row=row, column=col).fill = _SECTION_FILL


class QuantitiesWorkbookBuilder:
    """Monta a aba 'Quantities' linha a linha, guardando os ranges usados nos
    totais de categoria para depois alimentar a aba 'Resumo' por fórmula.
    """

    def __init__(self, building_name: str = "BUILDING 01"):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Quantities"
        self.building_name = building_name
        self._gib_total_cells: list[str] = []
        self._stop_total_cells: list[str] = []
        self._paint_total_cells: list[str] = []
        self._board_type_rows: dict[str, int] = {
            board_type: _GIB_TABLE_FIRST_ROW + i for i, board_type in enumerate(GIB_BOARD_TYPES)
        }
        self._row = 1

    def _gib_sale_ref(self, board_type: str) -> str:
        row = self._board_type_rows[board_type]
        return _sale_cell(row)

    # -- aba "Taxas": único lugar onde os preços são preenchidos --------------
    def add_rates_sheet(
        self,
        margin: float,
        gib_rates: list[GibBoardRate],
        painting_cost: float,
        stopping_wall_cost: float,
        stopping_ceiling_cost: float,
        board_width_m: float = 1.2,
        board_height_m: float = 2.4,
        corner_trim_cost: float = 0.0,
        sealant_cost: float = 0.0,
        skirting_paint_cost: float = 0.0,
        single_door_cost: float = 0.0,
        double_door_cost: float = 0.0,
    ) -> None:
        ws = self.wb.create_sheet("Taxas", 0)
        ws.cell(row=2, column=2, value="Taxas (preencha os valores de custo; a venda calcula sozinha)").font = Font(bold=True, size=13)

        ws.cell(row=4, column=2, value="Margem (%)").font = _BOLD
        margin_cell = ws.cell(row=4, column=3, value=margin)
        margin_cell.number_format = "0%"
        margin_cell.fill = _INPUT_FILL

        ws.cell(row=6, column=2, value="Largura da chapa (m)")
        w_cell = ws.cell(row=6, column=3, value=board_width_m)
        w_cell.number_format = "0.00"
        w_cell.fill = _INPUT_FILL
        ws.cell(row=7, column=2, value="Altura da chapa (m)")
        h_cell = ws.cell(row=7, column=3, value=board_height_m)
        h_cell.number_format = "0.00"
        h_cell.fill = _INPUT_FILL
        ws.cell(row=8, column=2, value="Área da chapa (m²)")
        ws.cell(row=8, column=3, value=f"={TAXAS_BOARD_WIDTH_CELL}*{TAXAS_BOARD_HEIGHT_CELL}").number_format = "0.00"

        headers = {
            TAXAS_COL_LABEL: "Item",
            TAXAS_COL_INSTALL: "Instalação (custo $/m²)",
            TAXAS_COL_BOARD_COST: "Custo da chapa ($)",
            TAXAS_COL_BOARD_COST_M2: "Custo chapa ($/m²)",
            TAXAS_COL_COST_TOTAL: "Custo total ($/m² ou $/un)",
            TAXAS_COL_SALE: "Venda ($/m² ou $/un)",
        }
        for col, text in headers.items():
            ws.cell(row=10, column=col, value=text)
        _style_header(ws, 10, list(headers.keys()))

        rates_by_type = {r.board_type: r for r in gib_rates}
        for board_type, row in self._board_type_rows.items():
            rate = rates_by_type.get(board_type, GibBoardRate(board_type, 0.0, 0.0))
            ws.cell(row=row, column=TAXAS_COL_LABEL, value=board_type)
            install_cell = ws.cell(row=row, column=TAXAS_COL_INSTALL, value=rate.install_cost_m2)
            install_cell.number_format = "0.00"
            install_cell.fill = _INPUT_FILL
            board_cell = ws.cell(row=row, column=TAXAS_COL_BOARD_COST, value=rate.board_cost)
            board_cell.number_format = "0.00"
            board_cell.fill = _INPUT_FILL
            ws.cell(
                row=row, column=TAXAS_COL_BOARD_COST_M2,
                value=f"={get_column_letter(TAXAS_COL_BOARD_COST)}{row}/{TAXAS_BOARD_AREA_CELL}",
            ).number_format = "0.00"
            ws.cell(
                row=row, column=TAXAS_COL_COST_TOTAL,
                value=(
                    f"={get_column_letter(TAXAS_COL_INSTALL)}{row}+"
                    f"{get_column_letter(TAXAS_COL_BOARD_COST_M2)}{row}"
                ),
            ).number_format = "0.00"
            ws.cell(
                row=row, column=TAXAS_COL_SALE,
                value=f"={get_column_letter(TAXAS_COL_COST_TOTAL)}{row}*(1+{TAXAS_MARGIN_CELL})",
            ).number_format = "0.00"

        flat_items = [
            (_PAINTING_ROW, "Pintura ($/m²)", painting_cost),
            (_STOPPING_WALL_ROW, "Stopping parede ($/m²)", stopping_wall_cost),
            (_STOPPING_CEILING_ROW, "Stopping teto - Square Stop ($/m²)", stopping_ceiling_cost),
            (_CORNER_ROW, "Cantoneira / Corner trim ($/m)", corner_trim_cost),
            (_SEALANT_ROW, "Selante / Sealant ($/m)", sealant_cost),
            (_SKIRTING_ROW, "Pintura de rodapé - Skirting ($/m)", skirting_paint_cost),
            (_SINGLE_DOOR_ROW, "Porta simples - Single door ($/porta)", single_door_cost),
            (_DOUBLE_DOOR_ROW, "Porta dupla - Double door ($/porta)", double_door_cost),
        ]
        for row, label, cost in flat_items:
            ws.cell(row=row, column=TAXAS_COL_LABEL, value=label)
            cost_cell = ws.cell(row=row, column=TAXAS_COL_COST_TOTAL, value=cost)
            cost_cell.number_format = "0.00"
            cost_cell.fill = _INPUT_FILL
            ws.cell(
                row=row, column=TAXAS_COL_SALE,
                value=f"={get_column_letter(TAXAS_COL_COST_TOTAL)}{row}*(1+{TAXAS_MARGIN_CELL})",
            ).number_format = "0.00"

        for col, width in (
            (TAXAS_COL_LABEL, 34), (TAXAS_COL_INSTALL, 15), (TAXAS_COL_BOARD_COST, 14),
            (TAXAS_COL_BOARD_COST_M2, 13), (TAXAS_COL_COST_TOTAL, 16), (TAXAS_COL_SALE, 16),
        ):
            ws.column_dimensions[get_column_letter(col)].width = width

    # -- cabeçalho da aba Quantities ------------------------------------------
    def start(self) -> None:
        ws = self.ws
        ws.cell(row=self._row, column=1, value=self.building_name).font = Font(bold=True, size=13)
        header_row = self._row + 1
        ws.cell(row=header_row, column=COL_GIB_RATE, value="GIB").font = _BOLD
        ws.cell(row=header_row, column=COL_STOP_RATE, value="Stopping").font = _BOLD
        ws.cell(row=header_row, column=COL_PAINT_RATE, value="Painting").font = _BOLD

        col_row = header_row + 1
        labels = {
            COL_DESC: "Description",
            COL_QTY: "Qty",
            COL_QTY_UNIT: "unit",
            COL_HEIGHT: "height",
            COL_TOTAL: "Total",
            COL_TOTAL_UNIT: "unit",
            COL_GIB_RATE: "Rate",
            COL_GIB_TOTAL: "Total",
            COL_STOP_RATE: "Rate",
            COL_STOP_TOTAL: "Total",
            COL_PAINT_RATE: "Rate",
            COL_PAINT_TOTAL: "Total",
            COL_ROW_TOTAL: "Row Total",
        }
        for col, text in labels.items():
            ws.cell(row=col_row, column=col, value=text)
        _style_header(ws, col_row, list(labels.keys()))

        self._row = col_row + 1

    def _write_total(self, r: int, qty_col: int = COL_QTY, mult_col: int = COL_HEIGHT) -> None:
        ws = self.ws
        ws.cell(
            row=r, column=COL_TOTAL,
            value=f"={get_column_letter(qty_col)}{r}*{get_column_letter(mult_col)}{r}",
        ).number_format = "0"

    def _write_row_total(self, r: int) -> None:
        ws = self.ws
        ws.cell(
            row=r, column=COL_ROW_TOTAL,
            value=(
                f"=SUM({get_column_letter(COL_GIB_TOTAL)}{r},"
                f"{get_column_letter(COL_STOP_TOTAL)}{r},"
                f"{get_column_letter(COL_PAINT_TOTAL)}{r})"
            ),
        ).number_format = "0.00"

    def _write_gib(self, r: int, rate_formula: str) -> None:
        ws = self.ws
        gib_rate_cell = f"{get_column_letter(COL_GIB_RATE)}{r}"
        ws.cell(row=r, column=COL_GIB_RATE, value=rate_formula).number_format = "0.00"
        ws.cell(
            row=r, column=COL_GIB_TOTAL,
            value=f"={get_column_letter(COL_TOTAL)}{r}*{gib_rate_cell}",
        ).number_format = "0.00"
        self._gib_total_cells.append(f"{get_column_letter(COL_GIB_TOTAL)}{r}")

    def _write_stopping(self, r: int, rate_formula: str) -> None:
        ws = self.ws
        stop_rate_cell = f"{get_column_letter(COL_STOP_RATE)}{r}"
        ws.cell(row=r, column=COL_STOP_RATE, value=rate_formula).number_format = "0.00"
        ws.cell(
            row=r, column=COL_STOP_TOTAL,
            value=f"={get_column_letter(COL_TOTAL)}{r}*{stop_rate_cell}",
        ).number_format = "0.00"
        self._stop_total_cells.append(f"{get_column_letter(COL_STOP_TOTAL)}{r}")

    def _write_painting(self, r: int, rate_formula: str) -> None:
        ws = self.ws
        paint_rate_cell = f"{get_column_letter(COL_PAINT_RATE)}{r}"
        ws.cell(row=r, column=COL_PAINT_RATE, value=rate_formula).number_format = "0.00"
        ws.cell(
            row=r, column=COL_PAINT_TOTAL,
            value=f"={get_column_letter(COL_TOTAL)}{r}*{paint_rate_cell}",
        ).number_format = "0.00"
        self._paint_total_cells.append(f"{get_column_letter(COL_PAINT_TOTAL)}{r}")

    # -- seção de forros de parede por altura ------------------------------
    def add_height_group(self, group: HeightGroup) -> None:
        ws = self.ws
        self._row += 1
        _section_row(ws, self._row, group.name)
        self._row += 1

        for item in group.items:
            r = self._row
            ws.cell(row=r, column=COL_DESC, value=item.description)
            ws.cell(row=r, column=COL_QTY, value=item.qty)
            ws.cell(row=r, column=COL_QTY_UNIT, value=item.qty_unit)
            ws.cell(row=r, column=COL_HEIGHT, value=group.height_m)
            self._write_total(r)
            ws.cell(row=r, column=COL_TOTAL_UNIT, value=item.total_unit)

            gib_ref = self._gib_sale_ref(item.board_type)
            gib_formula = f"={item.layers}*{gib_ref}" if item.layers != 1 else f"={gib_ref}"
            self._write_gib(r, gib_formula)
            if item.include_stopping:
                self._write_stopping(r, f"={TAXAS_STOPPING_WALL}")
            if item.include_painting:
                self._write_painting(r, f"={TAXAS_PAINTING}")
            self._write_row_total(r)
            self._row += 1

        if group.corner_trims_qty:
            r = self._row
            ws.cell(row=r, column=COL_DESC, value="Corner Trims")
            ws.cell(row=r, column=COL_QTY, value=group.corner_trims_qty)
            ws.cell(row=r, column=COL_QTY_UNIT, value="ea")
            ws.cell(row=r, column=COL_HEIGHT, value=group.height_m)
            self._write_total(r)
            ws.cell(row=r, column=COL_TOTAL_UNIT, value="m")
            self._write_stopping(r, f"={TAXAS_CORNER}")
            ws.cell(row=r, column=COL_ROW_TOTAL, value=f"={get_column_letter(COL_STOP_TOTAL)}{r}").number_format = "0.00"
            self._row += 1

        if group.sealant_qty:
            r = self._row
            ws.cell(row=r, column=COL_DESC, value="Sealant")
            ws.cell(row=r, column=COL_QTY, value=group.sealant_qty)
            ws.cell(row=r, column=COL_QTY_UNIT, value="m")
            ws.cell(row=r, column=COL_HEIGHT, value=2)
            self._write_total(r)
            ws.cell(row=r, column=COL_TOTAL_UNIT, value="m")
            self._write_gib(r, f"={TAXAS_SEALANT}")
            ws.cell(row=r, column=COL_ROW_TOTAL, value=f"={get_column_letter(COL_GIB_TOTAL)}{r}").number_format = "0.00"
            self._row += 1

    # -- seção de tetos -------------------------------------------------------
    def add_ceilings(self, items: list[CeilingItem], square_stop_qty: float = 0.0) -> None:
        ws = self.ws
        self._row += 1
        _section_row(ws, self._row, "CEILINGS")
        self._row += 1

        for item in items:
            r = self._row
            ws.cell(row=r, column=COL_DESC, value=item.description)
            ws.cell(row=r, column=COL_QTY, value=item.area_m2)
            ws.cell(row=r, column=COL_QTY_UNIT, value="m²")
            ws.cell(row=r, column=COL_HEIGHT, value=1)
            self._write_total(r)
            ws.cell(row=r, column=COL_TOTAL_UNIT, value="m2")

            self._write_gib(r, f"={self._gib_sale_ref(item.board_type)}")
            if item.include_stopping:
                self._write_stopping(r, f"={TAXAS_STOPPING_WALL}")
            if item.include_painting:
                self._write_painting(r, f"={TAXAS_PAINTING}")
            self._write_row_total(r)
            self._row += 1

        if square_stop_qty:
            r = self._row
            ws.cell(row=r, column=COL_DESC, value="Square Stop")
            ws.cell(row=r, column=COL_QTY, value=square_stop_qty)
            ws.cell(row=r, column=COL_QTY_UNIT, value="m")
            ws.cell(row=r, column=COL_HEIGHT, value=1)
            self._write_total(r)
            ws.cell(row=r, column=COL_TOTAL_UNIT, value="m")
            self._write_stopping(r, f"={TAXAS_STOPPING_CEILING}")
            ws.cell(row=r, column=COL_ROW_TOTAL, value=f"={get_column_letter(COL_STOP_TOTAL)}{r}").number_format = "0.00"
            self._row += 1

    # -- seção de pintura avulsa (rodapé, portas) ------------------------------
    def add_painting_only(self, skirting_m: float, doors: list[DoorPaintItem]) -> None:
        ws = self.ws
        self._row += 1
        _section_row(ws, self._row, "PAINTING")
        self._row += 1

        if skirting_m:
            r = self._row
            ws.cell(row=r, column=COL_DESC, value="60x10 Bevelled Edge Skirting")
            ws.cell(row=r, column=COL_QTY, value=skirting_m)
            ws.cell(row=r, column=COL_QTY_UNIT, value="m")
            ws.cell(row=r, column=COL_HEIGHT, value=1)
            self._write_total(r)
            ws.cell(row=r, column=COL_TOTAL_UNIT, value="m")
            self._write_painting(r, f"={TAXAS_SKIRTING_PAINT}")
            ws.cell(row=r, column=COL_ROW_TOTAL, value=f"={get_column_letter(COL_PAINT_TOTAL)}{r}").number_format = "0.00"
            self._row += 1

        for door in doors:
            r = self._row
            ws.cell(row=r, column=COL_DESC, value=door.description)
            ws.cell(row=r, column=COL_QTY, value=door.count)
            ws.cell(row=r, column=COL_QTY_UNIT, value="ea")
            ws.cell(row=r, column=COL_HEIGHT, value=1)
            self._write_total(r)
            ws.cell(row=r, column=COL_TOTAL_UNIT, value="each")
            rate_ref = TAXAS_SINGLE_DOOR if door.door_type == "single" else TAXAS_DOUBLE_DOOR
            self._write_painting(r, f"={rate_ref}")
            ws.cell(row=r, column=COL_ROW_TOTAL, value=f"={get_column_letter(COL_PAINT_TOTAL)}{r}").number_format = "0.00"
            self._row += 1

    # -- aba Resumo -----------------------------------------------------------
    def add_summary_sheet(self) -> None:
        ws = self.wb.create_sheet("Summary", 1)

        def joined(cells: list[str]) -> str:
            if not cells:
                return "=0"
            refs = ",".join(f"'Quantities'!{c}" for c in cells)
            return f"=SUM({refs})"

        ws.cell(row=2, column=4, value="TOTAL PROJECT").font = Font(bold=True, size=13)
        ws.cell(row=4, column=3, value="GIB").font = _BOLD
        ws.cell(row=4, column=4, value="PLASTER").font = _BOLD
        ws.cell(row=4, column=5, value="PAINTING").font = _BOLD
        ws.cell(row=4, column=6, value="TOTAL").font = _BOLD

        ws.cell(row=5, column=2, value="TOTAL").font = _BOLD
        ws.cell(row=5, column=3, value=joined(self._gib_total_cells)).number_format = "#,##0.00"
        ws.cell(row=5, column=4, value=joined(self._stop_total_cells)).number_format = "#,##0.00"
        ws.cell(row=5, column=5, value=joined(self._paint_total_cells)).number_format = "#,##0.00"
        ws.cell(row=5, column=6, value="=SUM(C5:E5)").number_format = "#,##0.00"

        ws.cell(row=8, column=2, value="GROSS MARGIN (%)")
        ws.cell(row=8, column=3, value=0.20).number_format = "0%"
        ws.cell(row=9, column=2, value="P&G ($)")
        ws.cell(row=9, column=3, value=0)
        ws.cell(row=10, column=2, value="CONTINGENCY (%)")
        ws.cell(row=10, column=3, value=0.05).number_format = "0%"
        ws.cell(row=12, column=2, value="PREÇO FINAL DE VENDA").font = _BOLD
        ws.cell(row=12, column=3, value="=F5*(1+C10)+C9").number_format = "#,##0.00"

        for col, width in ((1, 4), (2, 22), (3, 14), (4, 14), (5, 14), (6, 14)):
            ws.column_dimensions[get_column_letter(col)].width = width

    def finish(self) -> Workbook:
        for col, width in (
            (COL_DESC, 30), (COL_QTY, 9), (COL_QTY_UNIT, 7), (COL_HEIGHT, 9),
            (COL_TOTAL, 9), (COL_TOTAL_UNIT, 7), (COL_GIB_RATE, 9), (COL_GIB_TOTAL, 11),
            (COL_STOP_RATE, 9), (COL_STOP_TOTAL, 11), (COL_PAINT_RATE, 9),
            (COL_PAINT_TOTAL, 11), (COL_ROW_TOTAL, 12),
        ):
            self.ws.column_dimensions[get_column_letter(col)].width = width
        self.ws.freeze_panes = "A5"
        return self.wb
