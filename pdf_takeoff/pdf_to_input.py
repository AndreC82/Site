"""Extração automática, a partir de um PDF de planta arquitetônica (que siga a
convenção de códigos de sistema GIB comuns na Nova Zelândia), da planilha de
entrada (Taxas/Paredes/Tetos) já pré-preenchida — pronta pra você conferir e
ajustar no Excel antes de gerar o orçamento final.

Uso:  python -m pdf_takeoff.pdf_to_input planta.pdf entrada_preenchida.xlsx

Isto NÃO substitui a conferência humana — a extração geométrica de plantas
reais tem margem de erro (paredes densas, blocos de nota confundidos com
ambientes, etc.). Sempre confira as linhas marcadas como "auto" antes de
confiar no orçamento final. Use também o PDF de conferência gerado junto
(--review-pdf) para comparar visualmente com o desenho original.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field

import pymupdf as fitz
from openpyxl.styles import Font
from shapely.geometry import Point

from .calibration import auto_detect_scale, extract_scale_from_title_block
from .extract import PageContent, TextWord, extract_pdf
from .geometry import area_m2, build_room_polygons, exclude_exterior_areas, filter_title_block_noise, perimeter_m
from .gib_spec_extract import (
    BoardSpec,
    extract_ceiling_legend,
    extract_fire_wall_callouts,
    extract_stud_height_m,
    extract_wall_default,
)
from .input_template import generate_blank_input
from .wall_linings_plan import (
    detect_filled_wall_groups,
    extract_colored_wall_segments,
    extract_scale_m_per_unit as _wl_extract_scale,
    extract_wall_linings_color_key,
    find_placed_keynote_tags,
    match_segments_to_keynotes,
    parse_keynote_legend,
)

_WET_KEYWORDS = {"vinyl", "wc", "kitchen", "bath", "shower", "laundry", "sink", "servery"}
_LABEL_IGNORE_RE = re.compile(r"^[\d.,/\-x×ø]+$", re.IGNORECASE)
_MIN_ROOM_AREA_M2 = 2.5
_FIRE_MATCH_MAX_DIST_M = 5.0  # distância máxima entre o callout e o contorno do ambiente


@dataclass
class DetectedRoom:
    label: str
    page_number: int
    level: str
    polygon: object
    perimeter_m: float
    area_m2: float
    is_wet: bool
    fire_spec: BoardSpec | None = None
    ceiling_code: str | None = None
    ceiling_area_m2: float = 0.0
    ceiling_spec: BoardSpec | None = None


@dataclass
class AnalysisResult:
    rooms: list[DetectedRoom] = field(default_factory=list)
    height_by_level: dict[str, float] = field(default_factory=dict)
    warnings: list[str] = field(default_factory=list)
    scale_confident: bool = True
    # Níveis (ex.: "Lower", "Upper") em que não achamos a altura de pé-direito
    # na planta e por isso precisamos perguntar pro usuário.
    levels_missing_height: set[str] = field(default_factory=set)
    # Preenchido só quando a convenção de "ambiente + perímetro" (GB-code)
    # não achou nada e um método alternativo (wall-linings-plan colorido, ou
    # parede em preenchimento sólido) foi usado em vez disso -- cada linha já
    # pronta pra ir direto na aba Paredes: (grupo, altura_m, descrição,
    # tipo_de_chapa, camadas, metros_lineares).
    wall_rows: list[tuple[str, float, str, str, int, float]] = field(default_factory=list)
    method: str = "room-perimeter"


def _room_label(polygon, words: list[TextWord]) -> str:
    inside = [w for w in words if polygon.contains(Point(w.center))]
    candidates = [w for w in inside if not _LABEL_IGNORE_RE.match(w.text) and len(w.text) >= 3]
    if not candidates:
        return ""
    centroid = polygon.centroid
    candidates.sort(key=lambda w: Point(w.center).distance(centroid))
    return candidates[0].text


def _room_words(polygon, words: list[TextWord]) -> list[str]:
    return [w.text for w in words if polygon.contains(Point(w.center))]


def _page_scale(page_text: str, page: PageContent) -> tuple[float, bool]:
    scale = extract_scale_from_title_block(page_text)
    if scale is not None:
        return scale, True
    result = auto_detect_scale(page.segments, page.words)
    return result.scale_m_per_unit, result.confident


_TITLE_SCALE_RATIO_RE = re.compile(r"1\s*:\s*(\d+)(?:\s*@\s*A\d)?", re.IGNORECASE)
_MIN_WHOLE_BUILDING_SCALE_RATIO = 40  # abaixo disso (ex.: 1:20, 1:10) é prancha de detalhe, não de planta


def _looks_like_whole_building_plan(page_text: str) -> bool:
    """Uma prancha de planta de piso/teto/incêndio do edifício inteiro é
    sempre desenhada numa escala "grande" (tipicamente 1:50 a 1:200). Uma
    prancha de detalhe construtivo (junção de parede, perfil de rodapé etc.)
    usa escala bem mais "de perto" (1:1 a 1:20) e às vezes só menciona
    "Floor Plan" de passagem numa nota — sem esse filtro ela entraria como
    se fosse uma prancha de planta de verdade. Sem informação de escala
    (ex.: página de capa), tratamos como não sendo uma planta.
    """
    match = _TITLE_SCALE_RATIO_RE.search(page_text)
    if not match:
        return False
    return int(match.group(1)) >= _MIN_WHOLE_BUILDING_SCALE_RATIO


def _page_level(page_text: str) -> str:
    match = re.search(r"\b(Lower|Upper|Ground|First)\b\s+(?:Floor|Ceiling|Fire Control)\s+Plan", page_text, re.IGNORECASE)
    return match.group(1).title() if match else "Default"


def _reconstruct_rooms(page: PageContent, page_text: str) -> tuple[list, float]:
    scale, confident = _page_scale(page_text, page)
    polys = build_room_polygons(page.segments, min_area=_MIN_ROOM_AREA_M2 / scale**2)
    polys = filter_title_block_noise(polys, page.width, page.height, scale, page.words)
    polys = exclude_exterior_areas(polys, page.words)
    return polys, scale


def _fallback_height_for_page(
    text: str, page_no: int, result: AnalysisResult, height_overrides: dict[str, float]
) -> tuple[str, float | None]:
    """Mesma lógica de altura-por-nível usada no método principal (procura
    'FCL'/altura de pé-direito no texto da própria prancha; se não achar,
    usa o valor que o usuário informou na tela, ou marca o nível como
    pendente pra perguntar), mas com uma chave de nível que funciona mesmo
    sem o texto '(Lower|Upper|...) Floor Plan' -- as pranchas nos métodos
    alternativos abaixo nem sempre seguem essa convenção de título."""
    level = _page_level(text)
    if level == "Default":
        level = f"Página {page_no + 1}"
    if level not in result.height_by_level:
        height_m = extract_stud_height_m(text) or height_overrides.get(level)
        if height_m:
            result.height_by_level[level] = height_m
            result.levels_missing_height.discard(level)
        else:
            result.levels_missing_height.add(level)
    return level, result.height_by_level.get(level)


def _try_wall_linings_plan_fallback(
    doc: "fitz.Document", pages_text: list[str], result: AnalysisResult, height_overrides: dict[str, float],
    extra_text: str = "",
) -> bool:
    """Tenta a convenção "Wall Linings Plan" (trecho de parede = linha fina
    colorida, tipo/camadas confirmados por keynote próximo -- ver
    `wall_linings_plan.py`). Usada quando a convenção de ambiente+perímetro
    não achou nada; varre todas as páginas, porque não há como saber de
    antemão quais são as pranchas de parede nesse formato. Devolve True se
    achou pelo menos um trecho de parede em alguma página (e já preenche
    `result.wall_rows`/avisos nesse caso)."""
    found_any = False
    for i, text in enumerate(pages_text):
        scale = _wl_extract_scale(text)
        if scale is None:
            continue
        page = doc[i]
        segments = extract_colored_wall_segments(page)
        if not segments:
            continue
        legend = parse_keynote_legend(text + ("\n" + extra_text if extra_text else ""))
        color_key = extract_wall_linings_color_key(page)
        tags = find_placed_keynote_tags(page, legend)
        matched = match_segments_to_keynotes(segments, tags, legend, color_key, scale)
        if not matched:
            continue
        level, height = _fallback_height_for_page(text, i, result, height_overrides)
        totals: dict[tuple[str, int], float] = {}
        unmapped_len = 0.0
        for seg in matched:
            if seg.board_spec is None:
                unmapped_len += seg.length_m
                continue
            key = (seg.board_spec.board_type_label, seg.board_spec.layers)
            totals[key] = totals.get(key, 0.0) + seg.length_m
        for (board_label, layers), length_m in totals.items():
            if length_m <= 0:
                continue
            found_any = True
            result.wall_rows.append((
                f"{level} {height or 0:.2f}m", height or 0.0,
                f"Wall Linings Plan pág.{i + 1} - {board_label}",
                board_label, layers, round(length_m, 1),
            ))
        if unmapped_len > 0:
            found_any = True
            result.wall_rows.append((
                f"{level} {height or 0:.2f}m", height or 0.0,
                f"Wall Linings Plan pág.{i + 1} - cor não identificada (CONFERIR tipo de chapa)",
                "10mm Standard", 1, round(unmapped_len, 1),
            ))
    if found_any:
        result.method = "wall-linings-plan"
        result.warnings.append(
            "Não achei o padrão de 'ambiente + perímetro' nesta planta — usei o padrão de 'Wall Linings "
            "Plan' (linha colorida + keynote) em vez disso. Confira com atenção as linhas marcadas 'CONFERIR' "
            "e o tipo de chapa de cada linha antes de fechar o orçamento; teto não é detectado por este método."
        )
    return found_any


def _try_filled_rect_fallback(
    doc: "fitz.Document", pages: list[PageContent], pages_text: list[str], result: AnalysisResult,
    height_overrides: dict[str, float],
) -> bool:
    """Tenta a convenção de parede desenhada como retângulo de preenchimento
    sólido colorido (ver `detect_filled_wall_groups`). Diferente do método
    acima, aqui não há keynote nem legenda confiável de ler de forma
    genérica -- por isso cada grupo vira uma linha "CONFERIR" pedindo pra
    confirmar manualmente o tipo de chapa daquela cor. Ainda assim é melhor
    que devolver zero: dá o comprimento medido, só falta o nome do produto."""
    found_any = False
    for i, text in enumerate(pages_text):
        if not _looks_like_whole_building_plan(text):
            continue
        scale, confident = _page_scale(text, pages[i])
        if not scale:
            continue
        groups = detect_filled_wall_groups(doc[i], scale)
        if not groups:
            continue
        level, height = _fallback_height_for_page(text, i, result, height_overrides)
        for g in groups:
            found_any = True
            color_hex = "#%02X%02X%02X" % tuple(round(c * 255) for c in g.color)
            result.wall_rows.append((
                f"{level} {height or 0:.2f}m", height or 0.0,
                f"Preenchimento pág.{i + 1} - cor {color_hex}, {g.thickness_mm:.0f}mm (CONFERIR tipo de chapa)",
                f"{round(g.thickness_mm / 5) * 5}mm Standard", 1, g.length_m,
            ))
    if found_any:
        result.method = "filled-rect"
        result.warnings.append(
            "Não achei o padrão de 'ambiente + perímetro' nem de 'Wall Linings Plan' nesta planta — as paredes "
            "parecem desenhadas como preenchimento sólido colorido. Medi o comprimento de cada cor/espessura, "
            "mas NÃO consigo identificar o produto GIB de cada cor automaticamente nesse formato — confira e "
            "corrija a coluna 'Tipo de chapa' de cada linha marcada 'CONFERIR' antes de fechar o orçamento; "
            "teto também não é detectado por este método."
        )
    return found_any


def analyze_pdf(
    pdf_path: str,
    height_overrides: dict[str, float] | None = None,
    extra_text: str = "",
) -> AnalysisResult:
    """height_overrides: {"Lower": 2.7, "Upper": 2.55, ...} — usado quando a
    altura de pé-direito não foi encontrada na planta pra um nível (veja
    `AnalysisResult.levels_missing_height`); passe os valores que o usuário
    informou manualmente pra completar a análise sem precisar adivinhar.

    extra_text: texto que o usuário colou/digitou na tela (ex.: a legenda de
    GIB copiada da própria planta, quando ela não foi extraída/reconhecida
    automaticamente) — some ao texto de cada prancha antes de rodar os
    mesmos leitores de legenda já usados no resto do módulo, então funciona
    com qualquer formato que eles já entendam, sem precisar de código novo.
    """
    height_overrides = height_overrides or {}
    extra_text = extra_text or ""
    doc = fitz.open(pdf_path)
    pages = extract_pdf(pdf_path)
    pages_text = [doc[i].get_text() for i in range(len(doc))]
    full_text = "\n".join(pages_text) + ("\n" + extra_text if extra_text else "")

    result = AnalysisResult()

    ceiling_legend = extract_ceiling_legend(full_text)
    wall_default, wall_wet = extract_wall_default(full_text)
    if wall_default is None:
        wall_default = BoardSpec(thickness_mm=10, product="Standard", layers=1)
        result.warnings.append(
            "Não encontrei a nota 'Wall Linings' na planta — assumindo 10mm Standard como padrão."
        )
    if wall_wet is None:
        wall_wet = wall_default

    # -- paredes: reconstrói ambientes em cada prancha de planta de piso -----
    floor_plan_rooms: dict[str, list[DetectedRoom]] = {}
    for i, text in enumerate(pages_text):
        if not re.search(r"Floor Plan", text, re.IGNORECASE) or re.search(r"Floor Plan\s*Dimensions", text, re.IGNORECASE):
            continue
        if not _looks_like_whole_building_plan(text):
            continue
        level = _page_level(text)
        if level not in result.height_by_level:
            height_m = extract_stud_height_m(text) or height_overrides.get(level)
            if height_m:
                result.height_by_level[level] = height_m
                result.levels_missing_height.discard(level)
            else:
                result.levels_missing_height.add(level)

        polys, scale = _reconstruct_rooms(pages[i], text)
        if not scale:
            continue
        for poly in polys:
            words = pages[i].words
            label = _room_label(poly, words)
            if not label:
                continue
            room_words = {w.lower() for w in _room_words(poly, words)}
            is_wet = bool(room_words & _WET_KEYWORDS)
            room = DetectedRoom(
                label=label,
                page_number=i + 1,
                level=level,
                polygon=poly,
                perimeter_m=perimeter_m(poly, scale),
                area_m2=area_m2(poly, scale),
                is_wet=is_wet,
            )
            floor_plan_rooms.setdefault(level, []).append(room)

    if not floor_plan_rooms:
        if not _try_wall_linings_plan_fallback(doc, pages_text, result, height_overrides, extra_text):
            _try_filled_rect_fallback(doc, pages, pages_text, result, height_overrides)
        if not result.wall_rows:
            result.warnings.append(
                "Não consegui identificar nenhuma prancha de 'Floor Plan' com ambientes reconhecíveis, nem "
                "o padrão de 'Wall Linings Plan' colorido, nem parede em preenchimento sólido. Este PDF pode "
                "usar uma convenção de desenho ainda não suportada — envie-o para revisão."
            )

    # -- tetos: outras pranchas (plano de teto, incêndio) usam as MESMAS
    # coordenadas da planta de piso (mesmo edifício redesenhado) — em vez de
    # reconstruir ambientes de novo nelas (menos confiável, menos linhas
    # desenhadas), só verificamos se a posição do código/callout cai dentro
    # de um ambiente já reconstruído na planta de piso daquele nível. A área
    # do teto de um ambiente é a mesma área do seu piso (teto reto padrão).
    for i, text in enumerate(pages_text):
        if not re.search(r"Ceiling Plan", text, re.IGNORECASE) or not _looks_like_whole_building_plan(text):
            continue
        level = _page_level(text)
        rooms_this_level = floor_plan_rooms.get(level, [])
        if not rooms_this_level:
            continue
        for word in pages[i].words:
            code_match = word.text if re.match(r"^C[1-9]$", word.text) else None
            if not code_match or code_match not in ceiling_legend:
                continue
            point = Point(word.center)
            room = next((r for r in rooms_this_level if r.polygon.contains(point)), None)
            if room is None or room.ceiling_code is not None:
                continue
            room.ceiling_code = code_match
            room.ceiling_area_m2 = room.area_m2
            room.ceiling_spec = ceiling_legend[code_match]

    # -- combate a incêndio: acha paredes resistentes a fogo e sobrescreve ---
    for i, text in enumerate(pages_text):
        if not re.search(r"Fire Control Plan", text, re.IGNORECASE) or not _looks_like_whole_building_plan(text):
            continue
        level = _page_level(text)
        callouts = extract_fire_wall_callouts(doc[i])
        rooms_this_level = floor_plan_rooms.get(level, [])
        if not callouts or not rooms_this_level:
            continue
        scale, _ = _page_scale(text, pages[i])
        for callout in callouts:
            if callout.rect is None:
                continue
            callout_point = Point((callout.rect.x0 + callout.rect.x1) / 2, (callout.rect.y0 + callout.rect.y1) / 2)
            nearest_room = min(
                rooms_this_level, key=lambda r: r.polygon.exterior.distance(callout_point), default=None
            )
            if nearest_room is None:
                continue
            dist_m = nearest_room.polygon.exterior.distance(callout_point) * scale
            if dist_m > _FIRE_MATCH_MAX_DIST_M:
                continue
            nearest_room.fire_spec = callout.spec

    for rooms in floor_plan_rooms.values():
        result.rooms.extend(rooms)

    return result


def _wall_board_type(room: DetectedRoom, default_spec: BoardSpec, wet_spec: BoardSpec) -> BoardSpec:
    if room.fire_spec is not None:
        return room.fire_spec
    return wet_spec if room.is_wet else default_spec


def write_prefilled_input(result: AnalysisResult, output_path: str) -> None:
    generate_blank_input(output_path)
    from openpyxl import load_workbook

    wb = load_workbook(output_path)

    full_text_default = BoardSpec(thickness_mm=10, product="Standard", layers=1)
    wet_default = BoardSpec(thickness_mm=10, product="Aqualine", layers=1)

    paredes = wb["Paredes"]
    row = 5
    if result.rooms:
        for room in result.rooms:
            if room.perimeter_m <= 0:
                continue
            height = result.height_by_level.get(room.level) or next(iter(result.height_by_level.values()), 2.7)
            spec = _wall_board_type(room, full_text_default, wet_default)
            group_name = f"{room.level} {height:.2f}m"
            paredes.cell(row=row, column=1, value=group_name)
            paredes.cell(row=row, column=2, value=round(height, 2))
            paredes.cell(row=row, column=3, value=f"{room.label} ({room.level} pág.{room.page_number})")
            board_label = spec.board_type_label if spec.is_known_board_type else full_text_default.board_type_label
            paredes.cell(row=row, column=4, value=board_label)
            paredes.cell(row=row, column=5, value=spec.layers)
            paredes.cell(row=row, column=6, value=round(room.perimeter_m, 1))
            row += 1
    else:
        # Sem ambientes detectados (método principal não achou nada) -- usa
        # as linhas já prontas de um método alternativo (wall-linings-plan
        # ou preenchimento sólido), se algum tiver funcionado.
        for group_name, height, desc, board_label, layers, qty_m in result.wall_rows:
            paredes.cell(row=row, column=1, value=group_name)
            paredes.cell(row=row, column=2, value=round(height, 2) if height else None)
            paredes.cell(row=row, column=3, value=desc)
            paredes.cell(row=row, column=4, value=board_label)
            paredes.cell(row=row, column=5, value=layers)
            paredes.cell(row=row, column=6, value=qty_m)
            row += 1

    tetos = wb["Tetos"]
    row = 7
    for room in result.rooms:
        if not room.ceiling_spec or room.ceiling_area_m2 <= 0:
            continue
        board_label = room.ceiling_spec.board_type_label if room.ceiling_spec.is_known_board_type else "13mm Standard"
        tetos.cell(row=row, column=1, value=f"{room.label} ({room.ceiling_code}, pág.{room.page_number})")
        tetos.cell(row=row, column=2, value=board_label)
        tetos.cell(row=row, column=3, value=round(room.ceiling_area_m2, 1))
        row += 1

    if result.warnings:
        notes_ws = wb.create_sheet("Avisos da Extração")
        notes_ws.cell(row=1, column=1, value="Avisos gerados automaticamente na extração — confira antes de usar.").font = Font(bold=True)
        for i, warning in enumerate(result.warnings, start=3):
            notes_ws.cell(row=i, column=1, value=warning)
        notes_ws.column_dimensions["A"].width = 100

    wb.save(output_path)


def render_review_pdf(source_pdf_path: str, result: AnalysisResult, output_path: str) -> None:
    """Desenha, sobre a própria planta original, o contorno de cada ambiente
    detectado com a área de parede/teto e o tipo de chapa atribuído — pra
    conferir visualmente antes de confiar na planilha pré-preenchida.
    """
    doc = fitz.open(source_pdf_path)
    try:
        by_page: dict[int, list[DetectedRoom]] = {}
        for room in result.rooms:
            by_page.setdefault(room.page_number, []).append(room)

        for page_index in range(len(doc)):
            page_rooms = by_page.get(page_index + 1, [])
            if not page_rooms:
                continue
            page = doc[page_index]
            shape = page.new_shape()
            for room in page_rooms:
                if room.perimeter_m <= 0:
                    continue
                coords = list(room.polygon.exterior.coords)
                color = (0.85, 0.1, 0.1) if room.fire_spec else ((0.1, 0.4, 0.8) if room.is_wet else (0.2, 0.6, 0.2))
                shape.draw_polyline([fitz.Point(x, y) for x, y in coords])
                shape.finish(color=color, fill=color, fill_opacity=0.15, width=1.2, closePath=True)
            shape.commit()

            for room in page_rooms:
                if room.perimeter_m <= 0:
                    continue
                lines = [room.label or "(sem rótulo)"]
                lines.append(f"Parede: {room.perimeter_m:.1f} m linear")
                if room.fire_spec:
                    lines.append(f"FOGO: {room.fire_spec.board_type_label} ({room.fire_spec.layers}x)")
                elif room.is_wet:
                    lines.append("Área molhada -> Aqualine")
                if room.ceiling_code:
                    lines.append(f"Teto {room.ceiling_code}: {room.ceiling_area_m2:.1f} m²")
                text = "\n".join(lines)
                centroid = room.polygon.centroid
                origin = fitz.Point(centroid.x - 45, centroid.y - 6)
                page.insert_textbox(
                    fitz.Rect(origin.x, origin.y, origin.x + 140, origin.y + 14 * len(lines)),
                    text, fontsize=6.5, color=(0, 0, 0), fill=(1, 1, 1), fill_opacity=0.75, align=0,
                )

        doc.save(output_path)
    finally:
        doc.close()


def main() -> int:
    import argparse

    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("pdf", help="Caminho do PDF da planta.")
    parser.add_argument("output", help="Caminho da planilha de entrada pré-preenchida a gerar.")
    parser.add_argument(
        "--review-pdf", dest="review_pdf", default=None,
        help="Se informado, gera também um PDF com o que foi detectado desenhado sobre a planta.",
    )
    args = parser.parse_args()

    result = analyze_pdf(args.pdf)
    write_prefilled_input(result, args.output)
    if args.review_pdf:
        render_review_pdf(args.pdf, result, args.review_pdf)
        print(f"PDF de conferência gerado em: {args.review_pdf}")

    n_wall = sum(1 for r in result.rooms if r.perimeter_m > 0)
    n_ceiling = sum(1 for r in result.rooms if r.ceiling_area_m2 > 0)
    print(f"Planilha de entrada pré-preenchida gerada em: {args.output}")
    print(f"  {n_wall} ambiente(s) com parede detectada, {n_ceiling} com teto detectado.")
    if result.warnings:
        print("  Avisos:")
        for w in result.warnings:
            print("  -", w)
    print(
        "\nIMPORTANTE: confira as linhas antes de usar — extração automática de plantas "
        "reais tem margem de erro. Compare com o PDF original antes de gerar o orçamento final."
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
