"""Gera a planilha de orçamento (.xlsx) a partir das linhas de quantitativo."""

from __future__ import annotations

from collections import defaultdict

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .takeoff import QuantityLine

CATEGORY_LABELS = {
    "wall_paint": "Pintura - Paredes",
    "ceiling_paint": "Pintura - Tetos",
    "drywall": "Gesso - Drywall",
}

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_DETAIL_SHEET = "Detalhe por Ambiente"
_SUMMARY_SHEET = "Resumo"

_DETAIL_HEADERS = [
    "Ambiente ID",
    "Página",
    "Nome do Ambiente",
    "Categoria",
    "Código",
    "Descrição",
    "Camadas",
    "Área (m²)",
    "Observações",
    "Conferido (S/N)",
]


def _style_header(ws: Worksheet, ncols: int, row: int = 1) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autofit(ws: Worksheet, widths: list[int]) -> None:
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def _write_category_sheet(wb: Workbook, category: str, lines: list[QuantityLine]) -> None:
    ws = wb.create_sheet(CATEGORY_LABELS[category])
    headers = ["Código", "Descrição", "Ambiente", "Página", "Camadas", "Área (m²)", "Observações"]
    ws.append(headers)
    _style_header(ws, len(headers))
    for line in sorted(lines, key=lambda l: (l.code, l.room_id)):
        ws.append(
            [
                line.code,
                line.description,
                line.room_label or line.room_id,
                line.page_number,
                line.layers,
                round(line.area_m2, 2),
                line.note,
            ]
        )
    total_row = ws.max_row + 1
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    total_col = 6
    ws.cell(
        row=total_row,
        column=total_col,
        value=f"=SUM(F2:F{total_row - 1})",
    ).font = Font(bold=True)
    _autofit(ws, [14, 42, 20, 8, 9, 12, 45])


def _write_summary_sheet(wb: Workbook, lines: list[QuantityLine]) -> None:
    ws = wb.create_sheet(_SUMMARY_SHEET, 0)
    headers = ["Categoria", "Código", "Descrição", "Nº de Ambientes", "Área Total (m²)"]
    ws.append(headers)
    _style_header(ws, len(headers))

    grouped: dict[tuple[str, str], list[QuantityLine]] = defaultdict(list)
    for line in lines:
        grouped[(line.category, line.code)].append(line)

    grand_total_row_refs = []
    for (category, code), group in sorted(grouped.items(), key=lambda kv: (kv[0][0], kv[0][1])):
        total = sum(l.area_m2 for l in group)
        ws.append(
            [
                CATEGORY_LABELS[category],
                code,
                group[0].description,
                len({l.room_id for l in group}),
                round(total, 2),
            ]
        )
        grand_total_row_refs.append(ws.max_row)

    total_row = ws.max_row + 1
    ws.cell(row=total_row, column=1, value="TOTAL GERAL").font = Font(bold=True)
    ws.cell(
        row=total_row,
        column=5,
        value=f"=SUM(E2:E{total_row - 1})",
    ).font = Font(bold=True)
    _autofit(ws, [22, 14, 42, 16, 16])


def _write_detail_sheet(wb: Workbook, lines: list[QuantityLine]) -> None:
    ws = wb.create_sheet(_DETAIL_SHEET)
    ws.append(_DETAIL_HEADERS)
    _style_header(ws, len(_DETAIL_HEADERS))
    for line in sorted(lines, key=lambda l: (l.page_number, l.room_id, l.category, l.code)):
        ws.append(
            [
                line.room_id,
                line.page_number,
                line.room_label,
                CATEGORY_LABELS[line.category],
                line.code,
                line.description,
                line.layers,
                round(line.area_m2, 2),
                line.note,
                "",
            ]
        )
    ws.freeze_panes = "A2"
    _autofit(ws, [12, 8, 22, 20, 14, 42, 9, 12, 45, 14])


def export_workbook(lines: list[QuantityLine], output_path: str) -> None:
    """Escreve a planilha final: Resumo, uma aba por categoria e o Detalhe por Ambiente (editável).

    A aba "Detalhe por Ambiente" é a fonte de verdade: edite valores nela e
    use `reimport_workbook` para recalcular Resumo/abas por categoria sem
    reprocessar o PDF.
    """
    wb = Workbook()
    wb.remove(wb.active)

    _write_summary_sheet(wb, lines)
    for category in ("wall_paint", "ceiling_paint", "drywall"):
        cat_lines = [l for l in lines if l.category == category]
        if cat_lines:
            _write_category_sheet(wb, category, cat_lines)
    _write_detail_sheet(wb, lines)

    wb.save(output_path)


_CATEGORY_LABEL_TO_KEY = {v: k for k, v in CATEGORY_LABELS.items()}


def reimport_workbook(path: str) -> list[QuantityLine]:
    """Lê a aba 'Detalhe por Ambiente' de uma planilha (possivelmente editada pelo usuário)
    e reconstrói as linhas de quantitativo, para regerar Resumo/abas por categoria
    refletindo os ajustes manuais.
    """
    wb = load_workbook(path, data_only=True)
    if _DETAIL_SHEET not in wb.sheetnames:
        raise ValueError(
            f"A planilha não contém a aba '{_DETAIL_SHEET}'. "
            "Reimporte um arquivo gerado por este programa."
        )
    ws = wb[_DETAIL_SHEET]
    lines: list[QuantityLine] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        room_id, page_number, room_label, category_label, code, description, layers, area, note, _conferido = row
        if room_id is None or code is None:
            continue
        category = _CATEGORY_LABEL_TO_KEY.get(category_label, category_label)
        lines.append(
            QuantityLine(
                room_id=str(room_id),
                page_number=int(page_number) if page_number is not None else 0,
                room_label=room_label or "",
                category=category,
                code=str(code),
                description=description or "",
                layers=int(layers) if layers else 1,
                area_m2=float(area) if area is not None else 0.0,
                note=note or "",
            )
        )
    return lines
