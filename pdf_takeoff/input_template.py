"""Planilha de entrada: em vez de responder perguntas num terminal (frágil —
um Enter errado passa despercebido), o usuário preenche uma tabela normal do
Excel (com menus suspensos pra evitar erro de digitação) e o programa lê esse
arquivo direto para gerar a planilha de orçamento final.

Gerar o modelo em branco:  python -m pdf_takeoff.input_template gerar entrada.xlsx
Construir o orçamento:     python -m pdf_takeoff.input_template construir entrada.xlsx orcamento.xlsx
"""

from __future__ import annotations

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from .quantities_workbook import (
    CeilingItem,
    DoorPaintItem,
    GIB_BOARD_TYPES,
    GibBoardRate,
    HeightGroup,
    LiningItem,
    QuantitiesWorkbookBuilder,
)

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_EXAMPLE_FILL = PatternFill("solid", fgColor="F2F2F2")
_EXAMPLE_FONT = Font(italic=True, color="808080")
_BOLD = Font(bold=True)

_BLANK_ROWS = 40  # linhas em branco pré-formatadas em cada tabela, prontas pra preencher


def _style_header(ws: Worksheet, row: int, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _add_dropdown(ws: Worksheet, col_letter: str, first_row: int, last_row: int, options: list[str]) -> None:
    formula = '"' + ",".join(options) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{first_row}:{col_letter}{last_row}")


# ============================== GERAR MODELO ================================


def generate_blank_input(output_path: str) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    _build_taxas_sheet(wb)
    _build_paredes_sheet(wb)
    _build_tetos_sheet(wb)
    _build_pintura_sheet(wb)

    wb.save(output_path)


def _build_taxas_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Taxas")
    ws.cell(row=2, column=2, value="Preencha as taxas de custo (a venda calcula sozinha com a margem)").font = Font(bold=True, size=13)

    ws.cell(row=4, column=2, value="Margem (%)")
    ws.cell(row=4, column=3, value=0.25).number_format = "0%"

    ws.cell(row=6, column=2, value="Largura da chapa (m)")
    ws.cell(row=6, column=3, value=1.2)
    ws.cell(row=7, column=2, value="Altura da chapa (m)")
    ws.cell(row=7, column=3, value=2.4)

    headers = ["Tipo de chapa", "Instalação ($/m²)", "Custo da chapa ($)"]
    for i, text in enumerate(headers, start=2):
        ws.cell(row=10, column=i, value=text)
    _style_header(ws, 10, 4)

    for i, board_type in enumerate(GIB_BOARD_TYPES):
        row = 11 + i
        ws.cell(row=row, column=2, value=board_type)

    other_items = [
        "Pintura ($/m²)",
        "Stopping parede ($/m²)",
        "Stopping teto - Square Stop ($/m²)",
        "Cantoneira / Corner trim ($/m)",
        "Selante / Sealant ($/m)",
        "Pintura de rodapé - Skirting ($/m)",
        "Porta simples - Single door ($/porta)",
        "Porta dupla - Double door ($/porta)",
    ]
    ws.cell(row=20, column=2, value="Outros itens").font = _BOLD
    for i, label in enumerate(other_items):
        ws.cell(row=21 + i, column=2, value=label)

    for col, width in ((2, 34), (3, 16), (4, 16)):
        ws.column_dimensions[get_column_letter(col)].width = width


def _build_paredes_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Paredes")
    ws.cell(row=1, column=1, value="Uma linha por tipo de forro em cada altura de pé-direito diferente.").font = Font(italic=True)

    headers = ["Grupo (nome da altura)", "Altura pé-direito (m)", "Descrição", "Tipo de chapa", "Camadas (1 ou 2)", "Metros lineares"]
    for i, text in enumerate(headers, start=1):
        ws.cell(row=3, column=i, value=text)
    _style_header(ws, 3, len(headers))

    ws.cell(row=4, column=1, value="EXEMPLO 2.70m")
    ws.cell(row=4, column=2, value=2.70)
    ws.cell(row=4, column=3, value="1x 10mm Standard")
    ws.cell(row=4, column=4, value="10mm Standard")
    ws.cell(row=4, column=5, value=1)
    ws.cell(row=4, column=6, value=60)
    for col in range(1, 7):
        cell = ws.cell(row=4, column=col)
        cell.fill = _EXAMPLE_FILL
        cell.font = _EXAMPLE_FONT

    first_data_row = 5
    last_data_row = first_data_row + _BLANK_ROWS
    _add_dropdown(ws, "D", first_data_row, last_data_row, GIB_BOARD_TYPES)
    _add_dropdown(ws, "E", first_data_row, last_data_row, ["1", "2"])

    trims_row = last_data_row + 3
    ws.cell(row=trims_row, column=1, value="Cantoneiras e selante por grupo (opcional)").font = _BOLD
    trims_headers = ["Grupo (mesmo nome usado acima)", "Cantoneiras (qtde)", "Selante (m)"]
    for i, text in enumerate(trims_headers, start=1):
        ws.cell(row=trims_row + 1, column=i, value=text)
    _style_header(ws, trims_row + 1, len(trims_headers))
    ws.cell(row=trims_row + 2, column=1, value="EXEMPLO 2.70m")
    ws.cell(row=trims_row + 2, column=2, value=8)
    ws.cell(row=trims_row + 2, column=3, value=95)
    for col in range(1, 4):
        cell = ws.cell(row=trims_row + 2, column=col)
        cell.fill = _EXAMPLE_FILL
        cell.font = _EXAMPLE_FONT

    for col, width in ((1, 22), (2, 18), (3, 26), (4, 18), (5, 14), (6, 16)):
        ws.column_dimensions[get_column_letter(col)].width = width


def _build_tetos_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Tetos")
    ws.cell(row=1, column=1, value="Uma linha por tipo de chapa de teto.").font = Font(italic=True)

    ws.cell(row=3, column=1, value="Square Stop - junta de teto (m):").font = _BOLD
    ws.cell(row=3, column=2, value=0)

    headers = ["Descrição", "Tipo de chapa", "Área (m²)"]
    for i, text in enumerate(headers, start=1):
        ws.cell(row=5, column=i, value=text)
    _style_header(ws, 5, len(headers))

    ws.cell(row=6, column=1, value="EXEMPLO teto Standard")
    ws.cell(row=6, column=2, value="13mm Standard")
    ws.cell(row=6, column=3, value=180)
    for col in range(1, 4):
        cell = ws.cell(row=6, column=col)
        cell.fill = _EXAMPLE_FILL
        cell.font = _EXAMPLE_FONT

    first_data_row = 7
    last_data_row = first_data_row + _BLANK_ROWS
    _add_dropdown(ws, "B", first_data_row, last_data_row, GIB_BOARD_TYPES)

    for col, width in ((1, 30), (2, 18), (3, 14)):
        ws.column_dimensions[get_column_letter(col)].width = width


def _build_pintura_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Pintura Avulsa")
    ws.cell(row=1, column=1, value="Rodapé e portas (itens que só levam pintura, sem Gib/Stopping).").font = Font(italic=True)

    ws.cell(row=3, column=1, value="Rodapé - Skirting (m):").font = _BOLD
    ws.cell(row=3, column=2, value=0)

    headers = ["Descrição", "Quantidade", "Tipo (single/double)"]
    for i, text in enumerate(headers, start=1):
        ws.cell(row=5, column=i, value=text)
    _style_header(ws, 5, len(headers))

    ws.cell(row=6, column=1, value="EXEMPLO Portas quartos")
    ws.cell(row=6, column=2, value=6)
    ws.cell(row=6, column=3, value="single")
    for col in range(1, 4):
        cell = ws.cell(row=6, column=col)
        cell.fill = _EXAMPLE_FILL
        cell.font = _EXAMPLE_FONT

    first_data_row = 7
    last_data_row = first_data_row + _BLANK_ROWS
    _add_dropdown(ws, "C", first_data_row, last_data_row, ["single", "double"])

    for col, width in ((1, 30), (2, 14), (3, 20)):
        ws.column_dimensions[get_column_letter(col)].width = width


# ============================== LER E CONSTRUIR ==============================


def _rows(ws: Worksheet, start_row: int, ncols: int):
    for row in ws.iter_rows(min_row=start_row, max_col=ncols, values_only=True):
        if all(v is None or v == "" for v in row):
            continue
        first = str(row[0]).strip() if row[0] is not None else ""
        if first.upper().startswith("EXEMPLO"):
            continue
        yield row


def _read_rates(ws: Worksheet) -> dict:
    margin = ws["C4"].value or 0.0
    board_width = ws["C6"].value or 1.2
    board_height = ws["C7"].value or 2.4

    gib_rates = []
    for i, board_type in enumerate(GIB_BOARD_TYPES):
        row = 11 + i
        install = ws.cell(row=row, column=3).value or 0.0
        board_cost = ws.cell(row=row, column=4).value or 0.0
        gib_rates.append(GibBoardRate(board_type, install_cost_m2=float(install), board_cost=float(board_cost)))

    values = [ws.cell(row=21 + i, column=3).value or 0.0 for i in range(8)]
    (
        painting_cost,
        stopping_wall_cost,
        stopping_ceiling_cost,
        corner_trim_cost,
        sealant_cost,
        skirting_paint_cost,
        single_door_cost,
        double_door_cost,
    ) = [float(v) for v in values]

    return dict(
        margin=float(margin),
        gib_rates=gib_rates,
        painting_cost=painting_cost,
        stopping_wall_cost=stopping_wall_cost,
        stopping_ceiling_cost=stopping_ceiling_cost,
        board_width_m=float(board_width),
        board_height_m=float(board_height),
        corner_trim_cost=corner_trim_cost,
        sealant_cost=sealant_cost,
        skirting_paint_cost=skirting_paint_cost,
        single_door_cost=single_door_cost,
        double_door_cost=double_door_cost,
    )


def _read_height_groups(ws: Worksheet) -> list[HeightGroup]:
    groups: dict[str, HeightGroup] = {}
    order: list[str] = []
    for row in _rows(ws, 4, 6):
        group_name, height, desc, board_type, layers, qty = row
        if not group_name or not board_type or qty is None:
            continue
        group_name = str(group_name).strip()
        if group_name not in groups:
            groups[group_name] = HeightGroup(name=group_name, height_m=float(height or 0))
            order.append(group_name)
        layers_int = int(layers) if layers else 1
        item_desc = desc or f"{layers_int}x {board_type}"
        groups[group_name].items.append(
            LiningItem(str(item_desc), qty=float(qty), board_type=str(board_type), layers=layers_int)
        )

    # tabela de cantoneiras/selante por grupo: procura o cabeçalho "Cantoneiras (qtde)"
    trims_header_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=2).value == "Cantoneiras (qtde)":
            trims_header_row = r
            break
    if trims_header_row:
        for row in _rows(ws, trims_header_row + 1, 3):
            group_name, trims, sealant = row
            if not group_name or group_name not in groups:
                continue
            groups[str(group_name).strip()].corner_trims_qty = float(trims or 0)
            groups[str(group_name).strip()].sealant_qty = float(sealant or 0)

    return [groups[name] for name in order]


def _read_ceilings(ws: Worksheet) -> tuple[list[CeilingItem], float]:
    square_stop = ws["B3"].value or 0.0
    items = []
    for row in _rows(ws, 6, 3):
        desc, board_type, area = row
        if not board_type or area is None:
            continue
        items.append(CeilingItem(str(desc or board_type), area_m2=float(area), board_type=str(board_type)))
    return items, float(square_stop)


def _read_painting_only(ws: Worksheet) -> tuple[float, list[DoorPaintItem]]:
    skirting = ws["B3"].value or 0.0
    doors = []
    for row in _rows(ws, 6, 3):
        desc, count, door_type = row
        if not count:
            continue
        door_type = str(door_type).strip().lower() if door_type else "single"
        if door_type not in ("single", "double"):
            door_type = "single"
        doors.append(DoorPaintItem(str(desc or "Portas"), count=int(count), door_type=door_type))
    return float(skirting), doors


def build_from_input(input_path: str, output_path: str, building_name: str | None = None) -> None:
    wb_in = load_workbook(input_path, data_only=True)

    rates = _read_rates(wb_in["Taxas"])
    height_groups = _read_height_groups(wb_in["Paredes"])
    ceiling_items, square_stop_qty = _read_ceilings(wb_in["Tetos"])
    skirting_m, doors = _read_painting_only(wb_in["Pintura Avulsa"])

    builder = QuantitiesWorkbookBuilder(building_name=building_name or "BUILDING 01")
    builder.add_rates_sheet(**rates)
    builder.start()

    for group in height_groups:
        builder.add_height_group(group)
    if ceiling_items or square_stop_qty:
        builder.add_ceilings(ceiling_items, square_stop_qty=square_stop_qty)
    if skirting_m or doors:
        builder.add_painting_only(skirting_m, doors)

    builder.add_summary_sheet()
    wb_out = builder.finish()
    wb_out.save(output_path)


def main() -> int:
    import argparse

    parser = argparse.ArgumentParser(description=__doc__)
    sub = parser.add_subparsers(dest="command", required=True)

    p_gen = sub.add_parser("gerar", help="Gera a planilha de entrada em branco.")
    p_gen.add_argument("output", help="Caminho do arquivo .xlsx a criar.")

    p_build = sub.add_parser("construir", help="Lê uma planilha de entrada preenchida e gera o orçamento.")
    p_build.add_argument("input", help="Planilha de entrada preenchida.")
    p_build.add_argument("output", help="Caminho da planilha de orçamento a gerar.")
    p_build.add_argument("--nome", default=None, help="Nome do projeto/edifício.")

    args = parser.parse_args()
    if args.command == "gerar":
        generate_blank_input(args.output)
        print(f"Planilha de entrada gerada em: {args.output}")
    else:
        build_from_input(args.input, args.output, building_name=args.nome)
        print(f"Planilha de orçamento gerada em: {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
