from openpyxl import load_workbook

from pdf_takeoff.input_template import build_from_input, generate_blank_input


def _fill_minimal_input(path: str) -> None:
    wb = load_workbook(path)

    taxas = wb["Taxas"]
    taxas["C4"] = 0.25
    for i in range(8):
        row = 11 + i
        taxas.cell(row=row, column=3, value=8.0)
        taxas.cell(row=row, column=4, value=55.0)
    for i, v in enumerate([25.0, 11.0, 9.0, 10.0, 9.0, 15.0, 250.0, 450.0]):
        taxas.cell(row=21 + i, column=3, value=v)

    paredes = wb["Paredes"]
    paredes["A5"] = "2.70m"
    paredes["B5"] = 2.70
    paredes["C5"] = "1x 10mm Standard"
    paredes["D5"] = "10mm Standard"
    paredes["E5"] = 1
    paredes["F5"] = 60

    tetos = wb["Tetos"]
    tetos["A7"] = "Teto Standard"
    tetos["B7"] = "13mm Standard"
    tetos["C7"] = 180

    pintura = wb["Pintura Avulsa"]
    pintura["A7"] = "Portas quartos"
    pintura["B7"] = 6
    pintura["C7"] = "single"

    wb.save(path)


def test_generate_blank_input_has_expected_sheets(tmp_path):
    path = str(tmp_path / "entrada.xlsx")
    generate_blank_input(path)

    wb = load_workbook(path)
    assert set(wb.sheetnames) == {"Taxas", "Paredes", "Tetos", "Pintura Avulsa"}


def test_example_rows_are_not_read_as_real_data(tmp_path):
    """As linhas 'EXEMPLO...' do modelo em branco não devem virar quantidades."""
    input_path = str(tmp_path / "entrada.xlsx")
    output_path = str(tmp_path / "orcamento.xlsx")
    generate_blank_input(input_path)

    # Preenche só as taxas (mínimo pra não quebrar), deixa as tabelas de
    # quantidade só com as linhas de EXEMPLO que já vêm no modelo.
    wb = load_workbook(input_path)
    taxas = wb["Taxas"]
    taxas["C4"] = 0.25
    for i in range(8):
        taxas.cell(row=11 + i, column=3, value=8.0)
        taxas.cell(row=11 + i, column=4, value=55.0)
    for i, v in enumerate([25.0, 11.0, 9.0, 10.0, 9.0, 15.0, 250.0, 450.0]):
        taxas.cell(row=21 + i, column=3, value=v)
    wb.save(input_path)

    build_from_input(input_path, output_path)

    wb_out = load_workbook(output_path)
    ws = wb_out["Quantities"]
    data_rows = [
        row for row in ws.iter_rows(min_row=5, values_only=True) if any(v is not None for v in row)
    ]
    assert data_rows == []


def test_build_from_filled_input_produces_nonzero_quantities(tmp_path):
    input_path = str(tmp_path / "entrada.xlsx")
    output_path = str(tmp_path / "orcamento.xlsx")
    generate_blank_input(input_path)
    _fill_minimal_input(input_path)

    build_from_input(input_path, output_path, building_name="Teste")

    wb_out = load_workbook(output_path)
    assert set(wb_out.sheetnames) >= {"Taxas", "Summary", "Quantities"}

    ws = wb_out["Quantities"]
    data_rows = [
        row for row in ws.iter_rows(min_row=5, values_only=True) if any(v is not None for v in row)
    ]
    # 1 linha de parede + 1 linha de teto = pelo menos 2 linhas de dado
    assert len(data_rows) >= 2
