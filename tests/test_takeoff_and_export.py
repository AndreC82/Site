import pytest

from pdf_takeoff.export import export_workbook, reimport_workbook
from pdf_takeoff.extract import PageContent, TextWord
from pdf_takeoff.legend import LegendEntry
from pdf_takeoff.takeoff import build_takeoff

LEGEND = {
    "P-01": LegendEntry(code="P-01", category="wall_paint", description="Branco"),
    "PT-01": LegendEntry(code="PT-01", category="ceiling_paint", description="Teto branco"),
    "GB-AQ-DL": LegendEntry(
        code="GB-AQ-DL", category="drywall", description="Aqualine dupla", layers=2
    ),
}


def _rect_segments(x0, y0, x1, y1):
    return [
        ((x0, y0), (x1, y0)),
        ((x1, y0), (x1, y1)),
        ((x1, y1), (x0, y1)),
        ((x0, y1), (x0, y0)),
    ]


def _one_room_page():
    return PageContent(
        page_number=1,
        width=300,
        height=300,
        segments=_rect_segments(0, 0, 200, 150),
        words=[
            TextWord(text="P-01", bbox=(20, 20, 40, 30)),
            TextWord(text="PT-01", bbox=(20, 40, 45, 50)),
            TextWord(text="GB-AQ-DL", bbox=(20, 60, 60, 70)),
        ],
    )


def test_build_takeoff_computes_expected_areas():
    rooms, lines = build_takeoff([_one_room_page()], LEGEND, scale_m_per_unit=0.02, pe_direito_m=2.8)
    assert len(rooms) == 1
    room = rooms[0]

    # Tolerância folgada: o snap de vértices na reconstrução do polígono
    # desloca levemente os cantos em relação ao retângulo "ideal".
    expected_ceiling = 200 * 150 * 0.02**2
    expected_wall = 2 * (200 + 150) * 0.02 * 2.8
    assert room.ceiling_area_m2 == pytest.approx(expected_ceiling, abs=0.05)
    assert room.wall_area_m2 == pytest.approx(expected_wall, abs=0.05)

    by_category = {line.category: line for line in lines}
    assert by_category["wall_paint"].area_m2 == pytest.approx(expected_wall, abs=0.05)
    assert by_category["ceiling_paint"].area_m2 == pytest.approx(expected_ceiling, abs=0.05)
    # drywall de camada dupla: área de placa = área de parede x 2 camadas
    assert by_category["drywall"].area_m2 == pytest.approx(expected_wall * 2, abs=0.1)


def test_export_and_reimport_roundtrip(tmp_path):
    rooms, lines = build_takeoff([_one_room_page()], LEGEND, scale_m_per_unit=0.02, pe_direito_m=2.8)
    out = tmp_path / "orcamento.xlsx"
    export_workbook(lines, str(out))
    assert out.exists()

    reimported = reimport_workbook(str(out))
    assert len(reimported) == len(lines)
    total_before = sum(l.area_m2 for l in lines)
    total_after = sum(l.area_m2 for l in reimported)
    # export_workbook arredonda cada linha para 2 casas decimais; com N linhas
    # o erro acumulado máximo é N * 0.005.
    assert total_before == pytest.approx(total_after, abs=0.005 * len(lines) + 1e-9)


def test_reimport_reflects_manual_edits(tmp_path):
    rooms, lines = build_takeoff([_one_room_page()], LEGEND, scale_m_per_unit=0.02, pe_direito_m=2.8)
    out = tmp_path / "orcamento.xlsx"
    export_workbook(lines, str(out))

    from openpyxl import load_workbook

    wb = load_workbook(out)
    ws = wb["Detalhe por Ambiente"]
    # Usuário corrige manualmente a área da primeira linha de dados (linha 2)
    ws.cell(row=2, column=8, value=999.0)
    wb.save(out)

    reimported = reimport_workbook(str(out))
    edited = next(l for l in reimported if l.area_m2 == 999.0)
    assert edited is not None
