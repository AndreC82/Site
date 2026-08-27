"""Testes do casamento keynote<->linha colorida (pdf_takeoff.wall_linings_plan)
e do relatório de risco (pdf_takeoff.risk_report), usando uma prancha
sintética que reproduz o padrão real: linha colorida = tipo de chapa (ver
'WALL LININGS KEY'), tag de keynote perto = camadas/espessura exatas (ver
'Keynote Legend').
"""

from __future__ import annotations

import pytest
from reportlab.pdfgen import canvas

from pdf_takeoff.gib_spec_extract import BoardSpec
from pdf_takeoff.risk_report import (
    detect_duplicate_blocks,
    find_wall_segment_risks,
    write_bilingual_risk_report,
)
from pdf_takeoff.wall_linings_plan import (
    analyze_wall_linings_plan,
    parse_keynote_legend,
    summarize_by_board_spec,
)


def _draw_wall(c, x1, y, x2, color, width=1.4):
    c.setStrokeColorRGB(*color)
    c.setLineWidth(width)
    c.line(x1, y, x2, y)


def _draw_legend_swatch(c, x_text, y, color):
    c.setStrokeColorRGB(*color)
    c.setLineWidth(2.5)  # mais grossa que as paredes -> nao entra como segmento
    c.line(x_text - 40, y + 3, x_text - 10, y + 3)


def _generate_synthetic_plan(path: str) -> None:
    c = canvas.Canvas(path, pagesize=(1200, 900))
    c.setFont("Helvetica", 9)

    c.drawString(50, 850, "1:100")
    c.drawString(50, 830, "BUILDING TEST - GF WALL LININGS PLAN")

    # --- paredes desenhadas na planta (bem espacadas para o teste nao ter
    # falso-positivo de "o keynote vizinho e' de outra parede") ---
    # 1) azul (13mm Standard), com keynote "5113G 4.1" logo ao lado -> deve
    #    resolver para 1 camada, 13mm Standard, via keynote.
    _draw_wall(c, 100, 800, 220, color=(0, 0, 1))
    c.setFillColorRGB(0, 0, 0)
    c.drawString(160, 805, "5113G 4.1")

    # 2) magenta (13mm Fyreline) com keynote "5113G 4.5" (2 camadas) perto ->
    #    deve resolver para 2 camadas, nao 1, mesmo cor sendo so "fyreline".
    _draw_wall(c, 100, 600, 220, color=(1, 0, 1))
    c.drawString(160, 605, "5113G 4.5")

    # 3) magenta de novo, mas SEM nenhum keynote por perto (>3.5m de
    #    distancia de ambas as tags acima) -> deve cair pra "color_only"
    #    (1 camada assumida, sinalizado como risco).
    _draw_wall(c, 100, 400, 300, color=(1, 0, 1))

    # 4) verde, cor que nao existe na legenda -> "unmapped_color".
    _draw_wall(c, 100, 200, 180, color=(0, 1, 0))

    # --- legenda de cores (WALL LININGS KEY) ---
    c.drawString(700, 850, "13mm Gib Standard Plasterboard; Lvl 4 Paint Finish TBC")
    _draw_legend_swatch(c, 700, 850, (0, 0, 1))
    c.drawString(700, 835, "13mm Gib Fyreline Plasterboard; Lvl 4 Paint Finish TBC")
    _draw_legend_swatch(c, 700, 835, (1, 0, 1))

    # --- tabela Keynote Legend ---
    c.drawString(700, 160, "Keynote Legend")
    c.drawString(700, 145, "Key Value")
    c.drawString(700, 130, "5113G 4.1")
    c.drawString(700, 115, "ONE LAYER 13MM GIB STANDARD")
    c.drawString(700, 100, "5113G 4.5")
    c.drawString(700, 85, "TWO LAYERS 13MM GIB FYRELINE")

    c.showPage()
    c.save()


def test_parse_keynote_legend_reads_layers_and_thickness(tmp_path):
    text = (
        "Keynote Legend\n"
        "Key Value\n"
        "5113G 4.1\n"
        "ONE LAYER 13MM GIB STANDARD\n"
        "5113G 4.5\n"
        "TWO LAYERS 13MM GIB FYRELINE\n"
        "NOTE: something unrelated\n"
    )
    legend = parse_keynote_legend(text)
    assert legend["5113G 4.1"] == BoardSpec(thickness_mm=13, product="Standard", layers=1)
    assert legend["5113G 4.5"] == BoardSpec(thickness_mm=13, product="Fireline", layers=2)


def test_keynote_match_resolves_layers_that_color_alone_cannot(tmp_path):
    pdf_path = tmp_path / "plan.pdf"
    _generate_synthetic_plan(str(pdf_path))

    matched = analyze_wall_linings_plan(str(pdf_path))
    totals = summarize_by_board_spec(matched)

    standard_1l = BoardSpec(thickness_mm=13, product="Standard", layers=1)
    fireline_2l = BoardSpec(thickness_mm=13, product="Fireline", layers=2)

    assert standard_1l in totals
    assert fireline_2l in totals
    # o segmento (2) tem 2 camadas -- se o metodo so olhasse a cor, cairia
    # junto com o segmento (3) (1 camada assumida) num unico total errado.
    assert totals[fireline_2l] > 0

    sources = {m.source for m in matched}
    assert "keynote" in sources
    assert "color_only" in sources  # segmento (3), sem keynote por perto
    assert "unmapped_color" in sources  # segmento (4), cor verde nao mapeada


def test_risk_report_flags_color_only_and_unmapped_segments(tmp_path):
    pdf_path = tmp_path / "plan.pdf"
    _generate_synthetic_plan(str(pdf_path))
    matched = analyze_wall_linings_plan(str(pdf_path))

    findings = find_wall_segment_risks(matched, rate_lookup={"13mm Fireline": 29.55}, wall_height_m=3.0, sheet_label="GF")
    severities = {f.severity for f in findings}
    assert "medium" in severities  # color_only
    assert "high" in severities  # unmapped_color

    # o achado de "color_only" para o segmento (3) deve ter $ estimado, pois
    # demos uma taxa para 13mm Fireline.
    medium_findings = [f for f in findings if f.severity == "medium"]
    assert any(f.dollar_impact and f.dollar_impact > 0 for f in medium_findings)


def test_detect_duplicate_blocks_flags_identical_groups():
    groups = {
        "Building 1 - GF Ceiling": [137.15, 11.32, 115.76, 116.85],
        "Building 2 - GF Ceiling": [137.15, 11.32, 115.76, 116.85],  # copiado/colado
        "Building 2 - FF Ceiling": [206.03, 39.03, 29.19, 325.34],  # legitimamente diferente
    }
    findings = detect_duplicate_blocks(groups)
    assert len(findings) == 1
    assert "Building 1 - GF Ceiling" in findings[0].title_en
    assert "Building 2 - GF Ceiling" in findings[0].title_en
    assert findings[0].severity == "high"


def test_detect_duplicate_blocks_no_false_positive_on_different_groups():
    groups = {
        "A": [10.0, 20.0],
        "B": [10.0, 20.1],
    }
    assert detect_duplicate_blocks(groups) == []


def test_write_bilingual_risk_report_creates_both_languages(tmp_path):
    pdf_path = tmp_path / "plan.pdf"
    _generate_synthetic_plan(str(pdf_path))
    matched = analyze_wall_linings_plan(str(pdf_path))
    findings = find_wall_segment_risks(matched, sheet_label="GF")
    findings += detect_duplicate_blocks({"X": [1.0, 2.0], "Y": [1.0, 2.0]})

    out = tmp_path / "risk.xlsx"
    write_bilingual_risk_report(findings, str(out), quantities={"GF": summarize_by_board_spec(matched)})

    import openpyxl

    wb = openpyxl.load_workbook(out)
    assert "Findings - Achados" in wb.sheetnames
    assert "Quantities - Quantidades" in wb.sheetnames
    header = [c.value for c in next(wb["Findings - Achados"].iter_rows(min_row=4, max_row=4))]
    assert any("EN" in (h or "") for h in header)


def _generate_filled_wall_plan(path: str) -> None:
    """Prancha sintetica onde a parede e' um retangulo de preenchimento
    solido colorido (sem traco fino), convencao diferente da testada acima
    -- ex.: vermelho = 175mm, preto = 90mm, na escala 1:50 (mesma conversao
    usada no projeto real que motivou esse metodo). Inclui tambem um
    retangulo BRANCO (deve ser ignorado -- e' fundo/icone, nao parede) e um
    retangulo "quadrado" (deve ser ignorado -- proporcao baixa demais pra
    ser parede, ex. mobilia)."""
    scale = 50 * 25.4 / 1000 / 72  # m/pt, igual ao usado no resto do modulo
    red_th_pt = 0.175 / scale
    black_th_pt = 0.090 / scale

    c = canvas.Canvas(path, pagesize=(1200, 900))
    c.setFont("Helvetica", 9)
    c.drawString(50, 850, "1:50")
    c.drawString(50, 830, "LEVEL 9 - GENERAL ARRANGEMENT")  # sem "Floor Plan" de proposito

    c.setFillColorRGB(1, 0, 0)
    c.rect(100, 700, 400, red_th_pt, fill=1, stroke=0)  # parede vermelha ~70m (400pt)
    c.setFillColorRGB(0, 0, 0)
    c.rect(100, 500, 300, black_th_pt, fill=1, stroke=0)  # parede preta ~52.9m (300pt)
    c.setFillColorRGB(1, 1, 1)
    c.rect(100, 400, 400, red_th_pt, fill=1, stroke=0)  # branco -- deve ser ignorado
    c.setFillColorRGB(0, 1, 0)
    c.rect(100, 300, 30, 30, fill=1, stroke=0)  # quadrado verde -- deve ser ignorado (proporcao baixa)

    c.showPage()
    c.save()


def test_detect_filled_wall_groups_ignores_white_and_square_fills(tmp_path):
    import pymupdf as fitz

    from pdf_takeoff.wall_linings_plan import detect_filled_wall_groups

    pdf_path = tmp_path / "filled_plan.pdf"
    _generate_filled_wall_plan(str(pdf_path))

    doc = fitz.open(str(pdf_path))
    scale = 50 * 25.4 / 1000 / 72
    groups = detect_filled_wall_groups(doc[0], scale)

    colors = {g.color for g in groups}
    assert (1.0, 0.0, 0.0) in colors
    assert (0.0, 0.0, 0.0) in colors
    assert (1.0, 1.0, 1.0) not in colors  # branco filtrado
    assert (0.0, 1.0, 0.0) not in colors  # quadrado filtrado (proporcao)

    red = next(g for g in groups if g.color == (1.0, 0.0, 0.0))
    black = next(g for g in groups if g.color == (0.0, 0.0, 0.0))
    assert red.length_m == pytest.approx(400 * scale, rel=0.02)
    assert black.length_m == pytest.approx(300 * scale, rel=0.02)
    assert red.thickness_mm == pytest.approx(175, rel=0.05)
    assert black.thickness_mm == pytest.approx(90, rel=0.05)


def test_pdf_to_input_falls_back_to_wall_linings_plan(tmp_path):
    """Quando a convencao 'ambiente + perimetro' nao acha nada (esta
    prancha nao usa esse padrao), o pipeline principal (usado pelo webapp)
    deve cair automaticamente pro metodo de linha colorida + keynote, em vez
    de devolver zero -- e' o comportamento que corrige o bug relatado de
    'planilha gerada ficou zerada' para esse tipo de planta."""
    from pdf_takeoff.pdf_to_input import analyze_pdf

    pdf_path = tmp_path / "wl_plan.pdf"
    _generate_synthetic_plan(str(pdf_path))

    result = analyze_pdf(str(pdf_path), height_overrides={"Página 1": 2.7})

    assert result.method == "wall-linings-plan"
    assert len(result.rooms) == 0
    assert len(result.wall_rows) > 0
    board_labels = {row[3] for row in result.wall_rows}
    assert "13mm Standard" in board_labels
    assert "13mm Fireline" in board_labels


def test_pdf_to_input_falls_back_to_filled_rect(tmp_path):
    """Mesma ideia, mas para a convencao de parede em preenchimento solido
    (sem traco fino nem keynote) -- deve cair pro terceiro metodo, nao
    voltar zero."""
    from pdf_takeoff.pdf_to_input import analyze_pdf

    pdf_path = tmp_path / "filled_plan.pdf"
    _generate_filled_wall_plan(str(pdf_path))

    result = analyze_pdf(str(pdf_path), height_overrides={"Página 1": 2.7})

    assert result.method == "filled-rect"
    assert len(result.rooms) == 0
    assert len(result.wall_rows) == 2
    total_len = sum(row[5] for row in result.wall_rows)
    assert total_len == pytest.approx((400 + 300) * (50 * 25.4 / 1000 / 72), rel=0.05)
