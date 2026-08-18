from pdf_takeoff.wizard import run_wizard

ANSWERS = [
    "Teste",   # nome do projeto
    "custo", "25",  # taxas são custo, margem 25%
    "25",      # pintura
    "11",      # stopping parede
    "s", "9",  # separar stopping teto? sim, valor
    "s", "1.2", "2.4", "8", "55",  # gib único, dimensões chapa, instalação, custo chapa
    "8", "7", "15", "250", "450",  # cantoneira, selante, skirting, porta simples, porta dupla
    "s", "2.7", "",              # grupo de parede: altura, nome padrão
    "s", "1", "1", "60", "",     # item: 10mm Standard, 1 camada, 60m, descrição padrão
    "n",                          # não adicionar outro item
    "8", "95",                    # cantoneiras, selante do grupo
    "n",                          # não adicionar outro grupo
    "s", "3", "180", "",          # teto: 13mm Standard, 180m², descrição padrão
    "n",                          # não adicionar outra área de teto
    "220",                        # square stop
    "95",                         # skirting
    "s", "", "6", "single",       # portas: descrição padrão, 6, single
    "n",                          # não adicionar outro grupo de portas
]


def test_wizard_runs_end_to_end_and_produces_valid_workbook(tmp_path):
    answers = iter(ANSWERS)
    output = str(tmp_path / "wizard_teste.xlsx")

    path = run_wizard(input_fn=lambda _prompt: next(answers), output_path=output)

    assert path == output

    from openpyxl import load_workbook

    wb = load_workbook(output)
    assert set(wb.sheetnames) >= {"Taxas", "Summary", "Quantities"}
